from fastapi import FastAPI, Query
from fastapi.responses import HTMLResponse, StreamingResponse
import pandas as pd
import uvicorn
import threading
import webbrowser
import io

app = FastAPI()

# ---------------- HELPERS ----------------

def indian_format(n: float) -> str:
    """Format number in Indian style: 1,23,45,678.00"""
    n = round(float(n), 2)
    negative = n < 0
    n = abs(n)
    integer_part = int(n)
    decimal_part = f"{n - integer_part:.2f}"[1:]   # ".xx"

    s = str(integer_part)
    if len(s) <= 3:
        result = s
    else:
        # Last 3 digits, then groups of 2
        result = s[-3:]
        s = s[:-3]
        while s:
            result = s[-2:] + "," + result
            s = s[:-2]

    return ("-" if negative else "") + result + decimal_part

CY24_PATH = r"E:\Renault\Invoice Report CY24.csv"
CY25_PATH = r"E:\Renault\Invoice Report CY25.csv"


# ---------------- LOAD DATA ----------------

def load_data():
    df1 = pd.read_csv(CY24_PATH, low_memory=False)
    df2 = pd.read_csv(CY25_PATH, low_memory=False)

    df1["CY"] = "CY24"
    df2["CY"] = "CY25"

    df = pd.concat([df1, df2], ignore_index=True)
    df.columns = df.columns.str.strip()

    df["Division"] = df["Repair Order#"].astype(str).str[2:6]
    df["SA Name"]  = df["RO Owner First Name"].astype(str) + " " + df["RO Owner Last Name"].astype(str)

    # Format in CSV: "01-01-2025 10:06:00"  (DD-MM-YYYY HH:MM:SS)
    # Slice first 10 chars -> "01-01-2025", ignore time part
    for date_col in ["Invoice Date", "Repair Order Date"]:
        raw_date = df[date_col].astype(str).str.strip().str[:10]
        df[date_col] = pd.to_datetime(raw_date, format="%d-%m-%Y", errors="coerce")
        mask = df[date_col].isna()
        if mask.any():
            df.loc[mask, date_col] = pd.to_datetime(
                raw_date[mask], format="%d/%m/%Y", errors="coerce"
            )

    df["Month"] = df["Invoice Date"].dt.strftime("%b")  # Jan, Feb, Mar ...

    # TAT = Invoice Date - Repair Order Date (in days)
    df["TAT (Days)"] = (df["Invoice Date"] - df["Repair Order Date"]).dt.days

    for col in ["Net Taxable Labor Amount", "Net Taxable Parts Amt"]:
        df[col] = (
            df[col].astype(str)
            .str.replace("Rs.", "", regex=False)
            .str.replace(",",   "", regex=False)
            .pipe(pd.to_numeric, errors="coerce")
            .fillna(0)
        )

    return df


df = load_data()


# ---------------- FILTERS ----------------

@app.get("/filters")
def filters():
    return {
        "division": sorted(df["Division"].dropna().unique().tolist()),
        "service":  sorted(df["Service Type"].dropna().unique().tolist()),
        "sa":       sorted(df["SA Name"].dropna().unique().tolist()),
        "invoice":  sorted(df["Invoice Type"].dropna().unique().tolist()),
        "month":    sorted(
            df["Month"].dropna().unique().tolist(),
            key=lambda m: ["Jan","Feb","Mar","Apr","May","Jun",
                           "Jul","Aug","Sep","Oct","Nov","Dec"].index(m)
            if m in ["Jan","Feb","Mar","Apr","May","Jun",
                     "Jul","Aug","Sep","Oct","Nov","Dec"] else 99
        ),
    }


# ---------------- CARDS ----------------

@app.get("/cards")
def cards(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    month:    list[str] = Query(None),
    cy: str = None,
):
    data = df.copy()

    if cy:       data = data[data["CY"] == cy]
    if division: data = data[data["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]
    if month:    data = data[data["Month"].isin(month)]

    return {
        "inflow24": int(data[data["CY"] == "CY24"]["Repair Order#"].nunique()),
        "inflow25": int(data[data["CY"] == "CY25"]["Repair Order#"].nunique()),
        "labour":   "₹ " + indian_format(data['Net Taxable Labor Amount'].sum()),
        "spares":   "₹ " + indian_format(data['Net Taxable Parts Amt'].sum()),
    }


# ---------------- TABLE ----------------

@app.get("/table")
def table(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    month:    list[str] = Query(None),
    cy: str = None,
):
    data = df.copy()

    if cy:       data = data[data["CY"] == cy]
    if division: data = data[data["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]
    if month:    data = data[data["Month"].isin(month)]

    keep = ["Division", "Repair Order#", "SA Name", "Vehicle Reg#",
            "Service Type", "Net Taxable Labor Amount", "Net Taxable Parts Amt"]
    return data[keep].head(500).to_dict(orient="records")


# ---------------- EXPORT ----------------

@app.get("/export")
def export(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    month:    list[str] = Query(None),
    cy: str = None,
):
    data = df.copy()

    if cy:       data = data[data["CY"] == cy]
    if division: data = data[data["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]
    if month:    data = data[data["Month"].isin(month)]

    # All original columns except the computed ones we'll append at end
    computed = ["Division", "SA Name", "Month", "CY",
                "Net Taxable Labor Amount", "Net Taxable Parts Amt", "TAT (Days)"]
    base_cols = [c for c in data.columns if c not in computed]

    # Final column order: all originals first, then computed helpers, then Labour/Spares/TAT last
    export_cols = base_cols
    for col in ["Division", "SA Name", "Month", "CY"]:
        if col in data.columns:
            export_cols = export_cols + [col]
    # Labour, Spares, TAT always at the very end
    for col in ["Net Taxable Labor Amount", "Net Taxable Parts Amt", "TAT (Days)"]:
        if col in data.columns:
            export_cols = export_cols + [col]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        data[export_cols].to_excel(writer, index=False, sheet_name="Invoice Data")

        ws = writer.sheets["Invoice Data"]

        # Auto-fit column widths
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

        # Highlight last 3 columns (Labour, Spares, TAT) header in green
        from openpyxl.styles import PatternFill, Font
        green_fill = PatternFill("solid", fgColor="17A34A")
        white_font = Font(color="FFFFFF", bold=True)
        total_cols = ws.max_column
        for ci in range(total_cols - 2, total_cols + 1):
            cell = ws.cell(row=1, column=ci)
            cell.fill = green_fill
            cell.font = white_font

    output.seek(0)
    filename = "Renault_Invoice_Export.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ---------------- DASHBOARD ----------------

@app.get("/", response_class=HTMLResponse)
def dashboard():
    return HTMLResponse("""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>Renault Dashboard</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }

body {
    font-family: Arial, sans-serif;
    background: linear-gradient(135deg, #4f6bdc, #7b4fdc);
    min-height: 100vh;
    padding: 24px;
}

.container {
    background: #fff;
    border-radius: 16px;
    padding: 28px 32px;
    max-width: 1400px;
    margin: 0 auto;
}

h2 { font-size: 22px; margin-bottom: 20px; color: #222; }

/* ---------- CARDS ---------- */
.cards { display: flex; gap: 16px; margin-bottom: 24px; }
.card {
    flex: 1;
    background: #f4f6ff;
    border-radius: 10px;
    padding: 18px 16px;
    text-align: center;
}
.card .label {
    font-size: 12px; color: #777; font-weight: 600;
    text-transform: uppercase; letter-spacing: .5px; margin-bottom: 8px;
}
.card .value { font-size: 22px; font-weight: 700; color: #4f6bdc; }

/* ---------- FILTER BAR ---------- */
.filter-bar {
    display: flex;
    flex-wrap: wrap;
    gap: 12px;
    align-items: flex-end;
    background: #f7f8ff;
    border-radius: 10px;
    padding: 16px;
    margin-bottom: 20px;
}

.fg {
    display: flex;
    flex-direction: column;
    gap: 4px;
    flex: 1;
    min-width: 150px;
    position: relative;
}

.fg label {
    font-size: 11px;
    font-weight: 700;
    color: #555;
    text-transform: uppercase;
    letter-spacing: .5px;
}

/* CY native select — the reference style */
.fg select#cy {
    width: 100%;
    padding: 7px 10px;
    border: 1px solid #ccc;
    border-radius: 6px;
    font-size: 13px;
    background: #fff;
    cursor: pointer;
    color: #333;
    appearance: auto;
    height: 34px;
}
.fg select#cy:focus { outline: 2px solid #4f6bdc; border-color: transparent; }

/* ---------- CUSTOM MULTI-SELECT ---------- */
.custom-select { position: relative; width: 100%; }

/* The "face" button — identical look to #cy */
.cs-face {
    width: 100%;
    height: 34px;
    padding: 0 28px 0 10px;
    border: 1px solid #ccc;
    border-radius: 6px;
    font-size: 13px;
    background: #fff url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6'%3E%3Cpath d='M0 0l5 6 5-6z' fill='%23666'/%3E%3C/svg%3E") no-repeat right 10px center;
    cursor: pointer;
    text-align: left;
    color: #333;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    display: flex;
    align-items: center;
    user-select: none;
}
.cs-face:hover { border-color: #aaa; }
.cs-face.open   { outline: 2px solid #4f6bdc; border-color: transparent; }

/* Dropdown panel */
.cs-panel {
    display: none;
    position: absolute;
    top: calc(100% + 4px);
    left: 0;
    width: 100%;
    min-width: 200px;
    background: #fff;
    border: 1px solid #ccc;
    border-radius: 6px;
    box-shadow: 0 4px 16px rgba(0,0,0,.15);
    z-index: 999;
    max-height: 240px;
    overflow: hidden;
    flex-direction: column;
}
.cs-panel.open { display: flex; }

/* Search box inside panel */
.cs-search {
    padding: 8px 8px 4px;
    border-bottom: 1px solid #eee;
    flex-shrink: 0;
}
.cs-search input {
    width: 100%;
    padding: 5px 8px;
    border: 1px solid #ccc;
    border-radius: 4px;
    font-size: 12px;
    outline: none;
}
.cs-search input:focus { border-color: #4f6bdc; }

/* Options list */
.cs-list {
    overflow-y: auto;
    flex: 1;
    padding: 4px 0;
}
.cs-item {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 6px 10px;
    font-size: 13px;
    cursor: pointer;
    color: #333;
}
.cs-item:hover { background: #f0f3ff; }
.cs-item input[type=checkbox] { accent-color: #4f6bdc; cursor: pointer; }
.cs-item.checked { background: #eef1ff; font-weight: 600; }

/* Clear link */
.cs-footer {
    border-top: 1px solid #eee;
    padding: 5px 10px;
    display: flex;
    justify-content: flex-end;
    flex-shrink: 0;
}
.cs-clear {
    font-size: 11px;
    color: #4f6bdc;
    cursor: pointer;
    text-decoration: underline;
}

/* ---------- APPLY BUTTON ---------- */
.apply-btn {
    padding: 0 26px;
    height: 34px;
    background: #4f6bdc;
    color: #fff;
    border: none;
    border-radius: 6px;
    font-size: 14px;
    font-weight: 700;
    cursor: pointer;
    align-self: flex-end;
    white-space: nowrap;
}
.apply-btn:hover { background: #3a56c5; }

.export-btn {
    padding: 0 20px;
    height: 34px;
    background: #17a34a;
    color: #fff;
    border: none;
    border-radius: 6px;
    font-size: 14px;
    font-weight: 700;
    cursor: pointer;
    align-self: flex-end;
    white-space: nowrap;
}
.export-btn:hover { background: #15803d; }
.export-btn:disabled { background: #86c9a3; cursor: not-allowed; }

/* ---------- TABLE ---------- */
.table-wrap { overflow-x: auto; }
table { width: 100%; border-collapse: collapse; font-size: 13px; }
thead tr { background: #4f6bdc; color: #fff; }
th { padding: 10px 12px; text-align: left; white-space: nowrap; }
td { padding: 8px 12px; border-bottom: 1px solid #eee; white-space: nowrap; }
tbody tr:hover { background: #f4f6ff; }
.empty-msg { text-align: center; color: #999; padding: 24px; }
</style>
</head>

<body>
<div class="container">

  <h2>🚗 Renault Service Dashboard</h2>

  <!-- CARDS -->
  <div class="cards">
    <div class="card"><div class="label">Inflow CY24</div><div class="value" id="inflow24">…</div></div>
    <div class="card"><div class="label">Inflow CY25</div><div class="value" id="inflow25">…</div></div>
    <div class="card"><div class="label">Total Labour</div><div class="value" id="labour">…</div></div>
    <div class="card"><div class="label">Total Spares</div><div class="value" id="spares">…</div></div>
  </div>

  <!-- FILTER BAR -->
  <div class="filter-bar">

    <div class="fg">
      <label>Calendar Year</label>
      <select id="cy">
        <option value="">All CY</option>
        <option value="CY24">CY24</option>
        <option value="CY25">CY25</option>
      </select>
    </div>

    <div class="fg">
      <label>Division</label>
      <div class="custom-select" id="cs-division" data-key="division" data-placeholder="All Divisions"></div>
    </div>

    <div class="fg">
      <label>Service Type</label>
      <div class="custom-select" id="cs-service" data-key="service" data-placeholder="All Service Types"></div>
    </div>

    <div class="fg">
      <label>SA Name</label>
      <div class="custom-select" id="cs-sa" data-key="sa" data-placeholder="All SAs"></div>
    </div>

    <div class="fg">
      <label>Invoice Type</label>
      <div class="custom-select" id="cs-invoice" data-key="invoice" data-placeholder="All Invoice Types"></div>
    </div>

    <div class="fg">
      <label>Month</label>
      <div class="custom-select" id="cs-month" data-key="month" data-placeholder="All Months"></div>
    </div>

    <button class="apply-btn" onclick="applyFilters()">Apply</button>
    <button class="export-btn" onclick="exportExcel()">⬇ Export Excel</button>

  </div>

  <!-- TABLE -->
  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th>Division</th><th>RO #</th><th>SA Name</th>
          <th>Vehicle</th><th>Service Type</th>
          <th>Labour (₹)</th><th>Spares (₹)</th>
        </tr>
      </thead>
      <tbody id="tbody">
        <tr><td colspan="7" class="empty-msg">Loading…</td></tr>
      </tbody>
    </table>
  </div>

</div><!-- /container -->

<script>
/* ====================================================
   CUSTOM MULTI-SELECT COMPONENT
   ==================================================== */
const selections = {};   // key -> Set of selected values

function buildCustomSelect(container) {
    const key         = container.dataset.key;
    const placeholder = container.dataset.placeholder || "Select…";
    selections[key]   = new Set();

    container.innerHTML = `
      <div class="cs-face" id="face-${key}">${placeholder}</div>
      <div class="cs-panel" id="panel-${key}">
        <div class="cs-search"><input type="text" placeholder="Search…" id="search-${key}"></div>
        <div class="cs-list"  id="list-${key}"></div>
        <div class="cs-footer"><span class="cs-clear" onclick="clearSelect('${key}')">Clear all</span></div>
      </div>`;

    document.getElementById("face-" + key).addEventListener("click", (e) => {
        e.stopPropagation();
        togglePanel(key);
    });

    document.getElementById("search-" + key).addEventListener("input", (e) => {
        filterList(key, e.target.value.toLowerCase());
    });
}

function togglePanel(key) {
    // Close all other panels first
    document.querySelectorAll(".cs-panel.open").forEach(p => {
        if (p.id !== "panel-" + key) {
            p.classList.remove("open");
            document.getElementById("face-" + p.id.replace("panel-","")).classList.remove("open");
        }
    });
    const panel = document.getElementById("panel-" + key);
    const face  = document.getElementById("face-"  + key);
    const isOpen = panel.classList.toggle("open");
    face.classList.toggle("open", isOpen);
}

function fillCustomSelect(key, list) {
    const listEl = document.getElementById("list-" + key);
    listEl.innerHTML = "";
    list.forEach(v => {
        const item = document.createElement("div");
        item.className = "cs-item";
        item.dataset.value = v;
        item.innerHTML = `<input type="checkbox" value="${v}"> <span>${v}</span>`;
        item.querySelector("input").addEventListener("change", (e) => {
            if (e.target.checked) selections[key].add(v);
            else                  selections[key].delete(v);
            item.classList.toggle("checked", e.target.checked);
            updateFace(key);
        });
        listEl.appendChild(item);
    });
}

function filterList(key, query) {
    document.querySelectorAll(`#list-${key} .cs-item`).forEach(item => {
        item.style.display = item.dataset.value.toLowerCase().includes(query) ? "" : "none";
    });
}

function clearSelect(key) {
    selections[key].clear();
    document.querySelectorAll(`#list-${key} .cs-item`).forEach(item => {
        item.classList.remove("checked");
        item.querySelector("input").checked = false;
        item.style.display = "";
    });
    document.getElementById("search-" + key).value = "";
    updateFace(key);
}

function updateFace(key) {
    const face = document.getElementById("face-" + key);
    const placeholder = document.querySelector(`[data-key="${key}"]`).dataset.placeholder;
    const sel  = selections[key];
    if (sel.size === 0) {
        face.textContent = placeholder;
    } else if (sel.size === 1) {
        face.textContent = [...sel][0];
    } else {
        face.textContent = sel.size + " selected";
    }
}

// Close panels when clicking outside
document.addEventListener("click", () => {
    document.querySelectorAll(".cs-panel.open").forEach(p => {
        p.classList.remove("open");
        document.getElementById("face-" + p.id.replace("panel-","")).classList.remove("open");
    });
});

/* ====================================================
   API HELPERS
   ==================================================== */
function buildParams() {
    const p = new URLSearchParams();
    const cy = document.getElementById("cy").value;
    if (cy) p.append("cy", cy);
    ["division","service","sa","invoice","month"].forEach(key => {
        selections[key].forEach(v => p.append(key, v));
    });
    return p.toString();
}

function fmt(n) {
    return Number(n || 0).toLocaleString("en-IN", { minimumFractionDigits: 2, maximumFractionDigits: 2 });
}

async function loadFilters() {
    try {
        const data = await fetch("/filters").then(r => r.json());
        fillCustomSelect("division", data.division);
        fillCustomSelect("service",  data.service);
        fillCustomSelect("sa",       data.sa);
        fillCustomSelect("invoice",  data.invoice);
        fillCustomSelect("month",    data.month);
    } catch(e) { console.error("Filter load failed", e); }
}

async function loadCards() {
    try {
        const data = await fetch("/cards?" + buildParams()).then(r => r.json());
        document.getElementById("inflow24").innerText = Number(data.inflow24).toLocaleString("en-IN");
        document.getElementById("inflow25").innerText = Number(data.inflow25).toLocaleString("en-IN");
        document.getElementById("labour").innerText   = data.labour;
        document.getElementById("spares").innerText   = data.spares;
    } catch(e) { console.error("Card load failed", e); }
}

async function loadTable() {
    const tbody = document.getElementById("tbody");
    tbody.innerHTML = '<tr><td colspan="7" class="empty-msg">Loading…</td></tr>';
    try {
        const data = await fetch("/table?" + buildParams()).then(r => r.json());
        if (!data.length) {
            tbody.innerHTML = '<tr><td colspan="7" class="empty-msg">No records found</td></tr>';
            return;
        }
        tbody.innerHTML = data.map(r => `
          <tr>
            <td>${r.Division ?? ""}</td>
            <td>${r["Repair Order#"] ?? ""}</td>
            <td>${r["SA Name"] ?? ""}</td>
            <td>${r["Vehicle Reg#"] ?? ""}</td>
            <td>${r["Service Type"] ?? ""}</td>
            <td>${fmt(r["Net Taxable Labor Amount"])}</td>
            <td>${fmt(r["Net Taxable Parts Amt"])}</td>
          </tr>`).join("");
    } catch(e) {
        tbody.innerHTML = '<tr><td colspan="7" class="empty-msg">Error loading data</td></tr>';
        console.error("Table load failed", e);
    }
}

function applyFilters() { loadCards(); loadTable(); }

async function exportExcel() {
    const btn = document.querySelector(".export-btn");
    btn.disabled = true;
    btn.textContent = "⏳ Exporting…";
    try {
        const url = "/export?" + buildParams();
        const res = await fetch(url);
        if (!res.ok) throw new Error("Export failed");
        const blob = await res.blob();
        const a = document.createElement("a");
        a.href = URL.createObjectURL(blob);
        a.download = "Renault_Invoice_Export.xlsx";
        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
        URL.revokeObjectURL(a.href);
    } catch(e) {
        alert("Export failed: " + e.message);
        console.error(e);
    } finally {
        btn.disabled = false;
        btn.textContent = "⬇ Export Excel";
    }
}

/* ====================================================
   INIT
   ==================================================== */
window.onload = function () {
    // Build custom selects first (just the shell)
    document.querySelectorAll(".custom-select").forEach(buildCustomSelect);
    // Then load data
    loadFilters();
    loadCards();
    loadTable();
};
</script>
</body>
</html>""")


# ---------------- AUTO OPEN ----------------

def open_browser():
    webbrowser.open("http://127.0.0.1:8000")


if __name__ == "__main__":
    threading.Timer(1, open_browser).start()
    uvicorn.run(app, host="127.0.0.1", port=8000)
