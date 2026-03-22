from fastapi import FastAPI, Query
from fastapi.responses import HTMLResponse, StreamingResponse
import pandas as pd
import uvicorn
import threading
import webbrowser
import io

app = FastAPI()

# ---------------- HELPERS ----------------

def indian_format(n: float) -> str:
    """Format number in Indian style: 1,23,45,678.00"""
    n = round(float(n), 2)
    negative = n < 0
    n = abs(n)
    integer_part = int(n)
    decimal_part = f"{n - integer_part:.2f}"[1:]   # ".xx"

    s = str(integer_part)
    if len(s) <= 3:
        result = s
    else:
        # Last 3 digits, then groups of 2
        result = s[-3:]
        s = s[:-3]
        while s:
            result = s[-2:] + "," + result
            s = s[:-2]

    return ("-" if negative else "") + result + decimal_part

CY24_PATH = r"E:\Renault\Invoice Report CY24.csv"
CY25_PATH = r"E:\Renault\Invoice Report CY25.csv"


# ---------------- LOAD DATA ----------------

def load_data():
    df1 = pd.read_csv(CY24_PATH, low_memory=False)
    df2 = pd.read_csv(CY25_PATH, low_memory=False)

    df1["CY"] = "CY24"
    df2["CY"] = "CY25"

    df = pd.concat([df1, df2], ignore_index=True)
    df.columns = df.columns.str.strip()

    df["Division"] = df["Repair Order#"].astype(str).str[2:6]
    df["SA Name"]  = df["RO Owner First Name"].astype(str) + " " + df["RO Owner Last Name"].astype(str)
    df["Model"]    = df["Vehicle Model"].astype(str).str.strip().str.split().str[0].str.upper()

    # Format in CSV: "01-01-2025 10:06:00"  (DD-MM-YYYY HH:MM:SS)
    # Slice first 10 chars -> "01-01-2025", ignore time part
    for date_col in ["Invoice Date", "Repair Order Date"]:
        raw_date = df[date_col].astype(str).str.strip().str[:10]
        df[date_col] = pd.to_datetime(raw_date, format="%d-%m-%Y", errors="coerce")
        mask = df[date_col].isna()
        if mask.any():
            df.loc[mask, date_col] = pd.to_datetime(
                raw_date[mask], format="%d/%m/%Y", errors="coerce"
            )

    df["Month"] = df["Invoice Date"].dt.strftime("%b")  # Jan, Feb, Mar ...

    # TAT = Invoice Date - Repair Order Date (in days)
    df["TAT (Days)"] = (df["Invoice Date"] - df["Repair Order Date"]).dt.days

    for col in ["Net Taxable Labor Amount", "Net Taxable Parts Amt"]:
        df[col] = (
            df[col].astype(str)
            .str.replace("Rs.", "", regex=False)
            .str.replace(",",   "", regex=False)
            .pipe(pd.to_numeric, errors="coerce")
            .fillna(0)
        )

    return df


df = load_data()


# ---------------- FILTERS ----------------

@app.get("/filters")
def filters():
    return {
        "division": sorted(df["Division"].dropna().unique().tolist()),
        "service":  sorted(df["Service Type"].dropna().unique().tolist()),
        "sa":       sorted(df["SA Name"].dropna().unique().tolist()),
        "invoice":  sorted(df["Invoice Type"].dropna().unique().tolist()),
        "model":    sorted(df["Model"].dropna().unique().tolist()),
        "month":    sorted(
            df["Month"].dropna().unique().tolist(),
            key=lambda m: ["Jan","Feb","Mar","Apr","May","Jun",
                           "Jul","Aug","Sep","Oct","Nov","Dec"].index(m)
            if m in ["Jan","Feb","Mar","Apr","May","Jun",
                     "Jul","Aug","Sep","Oct","Nov","Dec"] else 99
        ),
    }


# ---------------- CARDS ----------------

@app.get("/cards")
def cards(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    month:    list[str] = Query(None),
    model:    list[str] = Query(None),
    cy: str = None,
):
    data = df.copy()

    if cy:       data = data[data["CY"] == cy]
    if division: data = data[data["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]
    if month:    data = data[data["Month"].isin(month)]
    if model:    data = data[data["Model"].isin(model)]

    return {
        "inflow24": int(data[data["CY"] == "CY24"]["Repair Order#"].nunique()),
        "inflow25": int(data[data["CY"] == "CY25"]["Repair Order#"].nunique()),
        "labour":   "₹ " + indian_format(data['Net Taxable Labor Amount'].sum()),
        "spares":   "₹ " + indian_format(data['Net Taxable Parts Amt'].sum()),
        "total":    "₹ " + indian_format(data['Net Taxable Labor Amount'].sum() + data['Net Taxable Parts Amt'].sum()),
    }


# ---------------- TABLE ----------------

@app.get("/table")
def table(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    month:    list[str] = Query(None),
    model:    list[str] = Query(None),
    cy: str = None,
):
    data = df.copy()

    if cy:       data = data[data["CY"] == cy]
    if division: data = data[data["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]
    if month:    data = data[data["Month"].isin(month)]
    if model:    data = data[data["Model"].isin(model)]

    keep = ["Division", "Repair Order#", "SA Name", "Vehicle Reg#",
            "Service Type", "Net Taxable Labor Amount", "Net Taxable Parts Amt"]
    out = data[keep].head(500).copy()
    out["Total Revenue"] = out["Net Taxable Labor Amount"] + out["Net Taxable Parts Amt"]
    return out.to_dict(orient="records")


# ---------------- EXPORT ----------------

@app.get("/export")
def export(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    month:    list[str] = Query(None),
    model:    list[str] = Query(None),
    cy: str = None,
):
    data = df.copy()

    if cy:       data = data[data["CY"] == cy]
    if division: data = data[data["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]
    if month:    data = data[data["Month"].isin(month)]
    if model:    data = data[data["Model"].isin(model)]

    # All original columns except the computed ones we'll append at end
    computed = ["Division", "SA Name", "Month", "CY",
                "Net Taxable Labor Amount", "Net Taxable Parts Amt", "TAT (Days)"]
    base_cols = [c for c in data.columns if c not in computed]

    # Final column order: all originals first, then computed helpers, then Labour/Spares/TAT last
    export_cols = base_cols
    for col in ["Division", "SA Name", "Month", "CY"]:
        if col in data.columns:
            export_cols = export_cols + [col]
    # Labour, Spares, TAT always at the very end
    for col in ["Net Taxable Labor Amount", "Net Taxable Parts Amt", "TAT (Days)"]:
        if col in data.columns:
            export_cols = export_cols + [col]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        data[export_cols].to_excel(writer, index=False, sheet_name="Invoice Data")

        ws = writer.sheets["Invoice Data"]

        # Auto-fit column widths
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

        # Highlight last 3 columns (Labour, Spares, TAT) header in green
        from openpyxl.styles import PatternFill, Font
        green_fill = PatternFill("solid", fgColor="17A34A")
        white_font = Font(color="FFFFFF", bold=True)
        total_cols = ws.max_column
        for ci in range(total_cols - 2, total_cols + 1):
            cell = ws.cell(row=1, column=ci)
            cell.fill = green_fill
            cell.font = white_font

    output.seek(0)
    filename = "Renault_Invoice_Export.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ---------------- COMPARISON ----------------

@app.get("/comparison")
def comparison(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    model:    list[str] = Query(None),
):
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    def apply_filters(data):
        if division: data = data[data["Division"].isin(division)]
        if service:  data = data[data["Service Type"].isin(service)]
        if sa:       data = data[data["SA Name"].isin(sa)]
        if invoice:  data = data[data["Invoice Type"].isin(invoice)]
        if model:    data = data[data["Model"].isin(model)]
        return data

    base24 = apply_filters(df[df["CY"] == "CY24"])
    base25 = apply_filters(df[df["CY"] == "CY25"])

    result = []
    for month in MONTHS:
        cy24 = base24[base24["Month"] == month]
        cy25 = base25[base25["Month"] == month]

        inflow24 = int(cy24["Repair Order#"].nunique())
        inflow25 = int(cy25["Repair Order#"].nunique())
        labour24 = float(cy24["Net Taxable Labor Amount"].sum())
        labour25 = float(cy25["Net Taxable Labor Amount"].sum())
        spares24 = float(cy24["Net Taxable Parts Amt"].sum())
        spares25 = float(cy25["Net Taxable Parts Amt"].sum())

        def pct(new, old):
            if old == 0: return None
            return round((new - old) / old * 100, 2)

        if inflow24 == 0 and inflow25 == 0:
            continue

        result.append({
            "month":      month,
            "inflow24":   inflow24,  "inflow25":   inflow25,  "inflow_pct": pct(inflow25, inflow24),
            "labour24":   labour24,  "labour25":   labour25,  "labour_pct": pct(labour25, labour24),
            "spares24":   spares24,  "spares25":   spares25,  "spares_pct": pct(spares25, spares24),
            "total24":    labour24 + spares24,
            "total25":    labour25 + spares25,
            "total_pct":  pct(labour25 + spares25, labour24 + spares24),
        })

    return result


# ---------------- DIVISION-MONTH ----------------

@app.get("/division-month")
def division_month(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    model:    list[str] = Query(None),
):
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    data = df.copy()
    if division: data = data[data["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]
    if model:    data = data[data["Model"].isin(model)]

    divisions = sorted(data["Division"].dropna().unique().tolist())

    # Build pivot: for each division, one row with values per month
    rows = []
    for div in divisions:
        div_data = data[data["Division"] == div]
        row = {"division": div}
        for month in MONTHS:
            m24 = div_data[(div_data["CY"] == "CY24") & (div_data["Month"] == month)]
            m25 = div_data[(div_data["CY"] == "CY25") & (div_data["Month"] == month)]
            row[f"i24_{month}"] = int(m24["Repair Order#"].nunique())
            row[f"i25_{month}"] = int(m25["Repair Order#"].nunique())
            row[f"l24_{month}"] = float(m24["Net Taxable Labor Amount"].sum())
            row[f"l25_{month}"] = float(m25["Net Taxable Labor Amount"].sum())
            row[f"s24_{month}"] = float(m24["Net Taxable Parts Amt"].sum())
            row[f"s25_{month}"] = float(m25["Net Taxable Parts Amt"].sum())
            row[f"t24_{month}"] = row[f"l24_{month}"] + row[f"s24_{month}"]
            row[f"t25_{month}"] = row[f"l25_{month}"] + row[f"s25_{month}"]
        rows.append(row)

    return {"divisions": divisions, "months": MONTHS, "rows": rows}

@app.get("/", response_class=HTMLResponse)
def dashboard():
    return HTMLResponse("""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>Renault Service Revenue Dashboard</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }

body {
    font-family: Arial, sans-serif;
    background: linear-gradient(135deg, #4f6bdc, #7b4fdc);
    min-height: 100vh;
    padding: 24px;
}

.container {
    background: #fff;
    border-radius: 16px;
    padding: 28px 32px;
    max-width: 1400px;
    margin: 0 auto;
}

/* ---------- HEADER ---------- */
.header { display: flex; align-items: center; justify-content: space-between; margin-bottom: 20px; }
.header h2 { font-size: 22px; color: #222; }

/* ---------- TABS ---------- */
.tabs { display: flex; gap: 4px; margin-bottom: 24px; border-bottom: 2px solid #e5e7eb; }
.tab-btn {
    padding: 10px 24px;
    border: none;
    background: none;
    font-size: 14px;
    font-weight: 600;
    color: #888;
    cursor: pointer;
    border-bottom: 3px solid transparent;
    margin-bottom: -2px;
    border-radius: 6px 6px 0 0;
    transition: all .2s;
}
.tab-btn:hover { color: #4f6bdc; background: #f4f6ff; }
.tab-btn.active { color: #4f6bdc; border-bottom-color: #4f6bdc; background: #f4f6ff; }

.tab-page { display: none; }
.tab-page.active { display: block; }

/* ---------- CARDS ---------- */
.cards { display: flex; gap: 16px; margin-bottom: 24px; }
.card {
    flex: 1; background: #f4f6ff; border-radius: 10px;
    padding: 18px 16px; text-align: center;
}
.card .label {
    font-size: 12px; color: #777; font-weight: 600;
    text-transform: uppercase; letter-spacing: .5px; margin-bottom: 8px;
}
.card .value { font-size: 22px; font-weight: 700; color: #4f6bdc; }

/* ---------- FILTER BAR ---------- */
.filter-bar {
    display: flex; flex-wrap: wrap; gap: 12px; align-items: flex-end;
    background: #f7f8ff; border-radius: 10px; padding: 16px; margin-bottom: 20px;
}
.fg { display: flex; flex-direction: column; gap: 4px; flex: 1; min-width: 150px; position: relative; }
.fg label {
    font-size: 11px; font-weight: 700; color: #555;
    text-transform: uppercase; letter-spacing: .5px;
}
.fg select#cy {
    width: 100%; padding: 7px 10px; border: 1px solid #ccc; border-radius: 6px;
    font-size: 13px; background: #fff; cursor: pointer; color: #333;
    appearance: auto; height: 34px;
}
.fg select#cy:focus { outline: 2px solid #4f6bdc; border-color: transparent; }

/* ---------- CUSTOM MULTI-SELECT ---------- */
.custom-select { position: relative; width: 100%; }
.cs-face {
    width: 100%; height: 34px; padding: 0 28px 0 10px; border: 1px solid #ccc;
    border-radius: 6px; font-size: 13px;
    background: #fff url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6'%3E%3Cpath d='M0 0l5 6 5-6z' fill='%23666'/%3E%3C/svg%3E") no-repeat right 10px center;
    cursor: pointer; text-align: left; color: #333; white-space: nowrap;
    overflow: hidden; text-overflow: ellipsis; display: flex; align-items: center; user-select: none;
}
.cs-face:hover { border-color: #aaa; }
.cs-face.open { outline: 2px solid #4f6bdc; border-color: transparent; }
.cs-panel {
    display: none; position: absolute; top: calc(100% + 4px); left: 0;
    width: 100%; min-width: 200px; background: #fff; border: 1px solid #ccc;
    border-radius: 6px; box-shadow: 0 4px 16px rgba(0,0,0,.15);
    z-index: 999; max-height: 240px; overflow: hidden; flex-direction: column;
}
.cs-panel.open { display: flex; }
.cs-search { padding: 8px 8px 4px; border-bottom: 1px solid #eee; flex-shrink: 0; }
.cs-search input {
    width: 100%; padding: 5px 8px; border: 1px solid #ccc;
    border-radius: 4px; font-size: 12px; outline: none;
}
.cs-search input:focus { border-color: #4f6bdc; }
.cs-list { overflow-y: auto; flex: 1; padding: 4px 0; }
.cs-item { display: flex; align-items: center; gap: 8px; padding: 6px 10px; font-size: 13px; cursor: pointer; color: #333; }
.cs-item:hover { background: #f0f3ff; }
.cs-item input[type=checkbox] { accent-color: #4f6bdc; cursor: pointer; }
.cs-item.checked { background: #eef1ff; font-weight: 600; }
.cs-footer { border-top: 1px solid #eee; padding: 5px 10px; display: flex; justify-content: flex-end; flex-shrink: 0; }
.cs-clear { font-size: 11px; color: #4f6bdc; cursor: pointer; text-decoration: underline; }

/* ---------- BUTTONS ---------- */
.apply-btn {
    padding: 0 26px; height: 34px; background: #4f6bdc; color: #fff;
    border: none; border-radius: 6px; font-size: 14px; font-weight: 700;
    cursor: pointer; align-self: flex-end; white-space: nowrap;
}
.apply-btn:hover { background: #3a56c5; }
.export-btn {
    padding: 0 20px; height: 34px; background: #17a34a; color: #fff;
    border: none; border-radius: 6px; font-size: 14px; font-weight: 700;
    cursor: pointer; align-self: flex-end; white-space: nowrap;
}
.export-btn:hover { background: #15803d; }
.export-btn:disabled { background: #86c9a3; cursor: not-allowed; }

/* ---------- TABLE (Page 1) ---------- */
.table-wrap { overflow-x: auto; }
table { width: 100%; border-collapse: collapse; font-size: 13px; }
thead tr { background: #4f6bdc; color: #fff; }
th { padding: 10px 12px; text-align: left; white-space: nowrap; }
td { padding: 8px 12px; border-bottom: 1px solid #eee; white-space: nowrap; }
tbody tr:hover { background: #f4f6ff; }
.empty-msg { text-align: center; color: #999; padding: 24px; }

/* ---------- COMPARISON TABLE (Page 2) ---------- */
.comp-section { margin-bottom: 36px; }
.comp-section h3 {
    font-size: 15px; font-weight: 700; color: #4f6bdc;
    margin-bottom: 12px; padding-bottom: 6px;
    border-bottom: 2px solid #e5e7eb;
}
.comp-table { width: 100%; border-collapse: collapse; font-size: 13px; }
.comp-table thead tr { background: #4f6bdc; color: #fff; }
.comp-table th { padding: 10px 14px; text-align: center; white-space: nowrap; }
.comp-table th:first-child { text-align: left; }
.comp-table td { padding: 9px 14px; border-bottom: 1px solid #eee; text-align: right; white-space: nowrap; }
.comp-table td:first-child { text-align: left; font-weight: 600; color: #333; }
.comp-table tbody tr:hover { background: #f4f6ff; }
.comp-table tfoot tr { background: #eef1ff; font-weight: 700; }
.comp-table tfoot td { padding: 10px 14px; border-top: 2px solid #4f6bdc; }

/* PCT badges */
.pct { display: inline-block; padding: 2px 8px; border-radius: 20px; font-size: 12px; font-weight: 700; }
.pct.up   { background: #dcfce7; color: #15803d; }
.pct.down { background: #fee2e2; color: #dc2626; }
.pct.flat { background: #f3f4f6; color: #6b7280; }

.comp-loading { text-align: center; color: #999; padding: 40px; font-size: 14px; }

/* ---------- DIVISION-MONTH PIVOT TABLE (Page 3) ---------- */
.pivot-wrap { overflow-x: auto; margin-bottom: 36px; }
.pivot-section h3 {
    font-size: 15px; font-weight: 700; color: #4f6bdc;
    margin-bottom: 12px; padding-bottom: 6px;
    border-bottom: 2px solid #e5e7eb;
}
.pivot-table { border-collapse: collapse; font-size: 11px; width: 100%; }
.pivot-table thead tr:first-child th {
    background: #4f6bdc; color: #fff; padding: 7px 8px;
    white-space: nowrap; text-align: center;
    border: 1px solid #3a56c5;
}
.pivot-table thead tr:first-child th:first-child { text-align: left; min-width: 80px; }
.pivot-table thead tr:nth-child(2) th {
    padding: 4px 6px; font-size: 10px; text-align: center;
    border: 1px solid #3a56c5; white-space: nowrap;
}
/* CY24 sub-header */
.pivot-table thead tr:nth-child(2) th:nth-child(3n+1) { background: #5a7ae8; color:#fff; }
/* CY25 sub-header */
.pivot-table thead tr:nth-child(2) th:nth-child(3n+2) { background: #4f6bdc; color:#fff; }
/* Growth% sub-header */
.pivot-table thead tr:nth-child(2) th:nth-child(3n)   { background: #2e4baa; color:#fff; }

.pivot-table td {
    padding: 6px 8px; border: 1px solid #e5e7eb;
    text-align: right; white-space: nowrap; font-size: 11px;
}
.pivot-table td:first-child {
    text-align: left; font-weight: 700; color: #4f6bdc;
    background: #f0f3ff; white-space: nowrap;
    position: sticky; left: 0; z-index: 1; min-width: 80px;
}
.pivot-table tbody tr:hover td            { background: #f4f6ff; }
.pivot-table tbody tr:hover td:first-child { background: #e8ecff; }
.pivot-table tfoot td {
    background: #dde3f8; font-weight: 700; color: #2a3e8c;
    border: 1px solid #b0bdee; border-top: 2px solid #4f6bdc;
    padding: 7px 8px; text-align: right;
}
.pivot-table tfoot td:first-child { text-align: left; position: sticky; left: 0; }
</style>
</head>

<body>
<div class="container">

  <!-- HEADER -->
  <div class="header">
    <h2>🚗 Renault Service Revenue Dashboard</h2>
  </div>

  <!-- TABS -->
  <div class="tabs">
    <button class="tab-btn active" onclick="switchTab('page1', this)">📋 Invoice Data</button>
    <button class="tab-btn"        onclick="switchTab('page2', this)">📊 CY Comparison</button>
    <button class="tab-btn"        onclick="switchTab('page3', this)">📅 Division-Month</button>
  </div>

  <!-- ==================== PAGE 1 ==================== -->
  <div id="page1" class="tab-page active">

    <!-- CARDS -->
    <div class="cards">
      <div class="card"><div class="label">Inflow CY24</div><div class="value" id="inflow24">…</div></div>
      <div class="card"><div class="label">Inflow CY25</div><div class="value" id="inflow25">…</div></div>
      <div class="card"><div class="label">Total Labour</div><div class="value" id="labour">…</div></div>
      <div class="card"><div class="label">Total Spares</div><div class="value" id="spares">…</div></div>
      <div class="card" style="border: 2px solid #4f6bdc;"><div class="label">Total Revenue</div><div class="value" id="total-rev">…</div></div>
    </div>

    <!-- FILTER BAR -->
    <div class="filter-bar">
      <div class="fg">
        <label>Calendar Year</label>
        <select id="cy">
          <option value="">All CY</option>
          <option value="CY24">CY24</option>
          <option value="CY25">CY25</option>
        </select>
      </div>
      <div class="fg"><label>Division</label>
        <div class="custom-select" id="cs-division" data-key="division" data-placeholder="All Divisions"></div>
      </div>
      <div class="fg"><label>Service Type</label>
        <div class="custom-select" id="cs-service" data-key="service" data-placeholder="All Service Types"></div>
      </div>
      <div class="fg"><label>SA Name</label>
        <div class="custom-select" id="cs-sa" data-key="sa" data-placeholder="All SAs"></div>
      </div>
      <div class="fg"><label>Invoice Type</label>
        <div class="custom-select" id="cs-invoice" data-key="invoice" data-placeholder="All Invoice Types"></div>
      </div>
      <div class="fg"><label>Model</label>
        <div class="custom-select" id="cs-model" data-key="model" data-placeholder="All Models"></div>
      </div>
      <div class="fg"><label>Month</label>
        <div class="custom-select" id="cs-month" data-key="month" data-placeholder="All Months"></div>
      </div>
      <button class="apply-btn"  onclick="applyFilters()">Apply</button>
      <button class="export-btn" onclick="exportExcel()">⬇ Export Excel</button>
    </div>

    <!-- TABLE -->
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Division</th><th>RO #</th><th>SA Name</th>
            <th>Vehicle</th><th>Service Type</th>
            <th>Labour (₹)</th><th>Spares (₹)</th>
            <th style="background:#3a56c5;">Total Revenue (₹)</th>
          </tr>
        </thead>
        <tbody id="tbody">
          <tr><td colspan="8" class="empty-msg">Loading…</td></tr>
        </tbody>
      </table>
    </div>

  </div><!-- /page1 -->

  <!-- ==================== PAGE 2 ==================== -->
  <div id="page2" class="tab-page">

    <!-- FILTER BAR (Comparison) -->
    <div class="filter-bar">
      <div class="fg">
        <label>Division</label>
        <div class="custom-select" id="cs2-division" data-key="c-division" data-placeholder="All Divisions"></div>
      </div>
      <div class="fg">
        <label>Service Type</label>
        <div class="custom-select" id="cs2-service" data-key="c-service" data-placeholder="All Service Types"></div>
      </div>
      <div class="fg">
        <label>SA Name</label>
        <div class="custom-select" id="cs2-sa" data-key="c-sa" data-placeholder="All SAs"></div>
      </div>
      <div class="fg">
        <label>Invoice Type</label>
        <div class="custom-select" id="cs2-invoice" data-key="c-invoice" data-placeholder="All Invoice Types"></div>
      </div>
      <div class="fg">
        <label>Model</label>
        <div class="custom-select" id="cs2-model" data-key="c-model" data-placeholder="All Models"></div>
      </div>
      <button class="apply-btn" onclick="applyComparison()">Apply</button>
    </div>

    <div id="comp-content">
      <div class="comp-loading">Click the tab to load comparison data…</div>
    </div>

  </div><!-- /page2 -->

  <!-- ==================== PAGE 3 ==================== -->
  <div id="page3" class="tab-page">

    <!-- FILTER BAR (Division-Month) -->
    <div class="filter-bar">
      <div class="fg">
        <label>Division</label>
        <div class="custom-select" id="cs3-division" data-key="d-division" data-placeholder="All Divisions"></div>
      </div>
      <div class="fg">
        <label>Service Type</label>
        <div class="custom-select" id="cs3-service" data-key="d-service" data-placeholder="All Service Types"></div>
      </div>
      <div class="fg">
        <label>SA Name</label>
        <div class="custom-select" id="cs3-sa" data-key="d-sa" data-placeholder="All SAs"></div>
      </div>
      <div class="fg">
        <label>Invoice Type</label>
        <div class="custom-select" id="cs3-invoice" data-key="d-invoice" data-placeholder="All Invoice Types"></div>
      </div>
      <div class="fg">
        <label>Model</label>
        <div class="custom-select" id="cs3-model" data-key="d-model" data-placeholder="All Models"></div>
      </div>
      <button class="apply-btn" onclick="applyDivMonth()">Apply</button>
    </div>

    <div id="divmonth-content">
      <div class="comp-loading">Click the tab to load data…</div>
    </div>

  </div><!-- /page3 -->

</div><!-- /container -->

<script>
/* ====================================================
   TAB SWITCHING
   ==================================================== */
let compLoaded    = false;
let divMonLoaded  = false;

function switchTab(pageId, btn) {
    document.querySelectorAll(".tab-page").forEach(p => p.classList.remove("active"));
    document.querySelectorAll(".tab-btn").forEach(b => b.classList.remove("active"));
    document.getElementById(pageId).classList.add("active");
    btn.classList.add("active");
    if (pageId === "page2" && !compLoaded) { loadComparison(); compLoaded = true; }
    if (pageId === "page3" && !divMonLoaded) { loadDivMonth(); divMonLoaded = true; }
}

/* ====================================================
   CUSTOM MULTI-SELECT COMPONENT
   ==================================================== */
const selections = {};

function buildCustomSelect(container) {
    const key         = container.dataset.key;
    const placeholder = container.dataset.placeholder || "Select…";
    selections[key]   = new Set();

    container.innerHTML = `
      <div class="cs-face" id="face-${key}">${placeholder}</div>
      <div class="cs-panel" id="panel-${key}">
        <div class="cs-search"><input type="text" placeholder="Search…" id="search-${key}"></div>
        <div class="cs-list"  id="list-${key}"></div>
        <div class="cs-footer"><span class="cs-clear" onclick="clearSelect('${key}')">Clear all</span></div>
      </div>`;

    document.getElementById("face-" + key).addEventListener("click", (e) => {
        e.stopPropagation(); togglePanel(key);
    });
    document.getElementById("search-" + key).addEventListener("input", (e) => {
        filterList(key, e.target.value.toLowerCase());
    });
}

function togglePanel(key) {
    document.querySelectorAll(".cs-panel.open").forEach(p => {
        if (p.id !== "panel-" + key) {
            p.classList.remove("open");
            document.getElementById("face-" + p.id.replace("panel-","")).classList.remove("open");
        }
    });
    const panel = document.getElementById("panel-" + key);
    const face  = document.getElementById("face-"  + key);
    panel.classList.toggle("open");
    face.classList.toggle("open", panel.classList.contains("open"));
}

function fillCustomSelect(key, list) {
    const listEl = document.getElementById("list-" + key);
    listEl.innerHTML = "";
    list.forEach(v => {
        const item = document.createElement("div");
        item.className = "cs-item";
        item.dataset.value = v;
        item.innerHTML = `<input type="checkbox" value="${v}"> <span>${v}</span>`;
        item.querySelector("input").addEventListener("change", (e) => {
            if (e.target.checked) selections[key].add(v);
            else                  selections[key].delete(v);
            item.classList.toggle("checked", e.target.checked);
            updateFace(key);
        });
        listEl.appendChild(item);
    });
}

function filterList(key, query) {
    document.querySelectorAll(`#list-${key} .cs-item`).forEach(item => {
        item.style.display = item.dataset.value.toLowerCase().includes(query) ? "" : "none";
    });
}

function clearSelect(key) {
    selections[key].clear();
    document.querySelectorAll(`#list-${key} .cs-item`).forEach(item => {
        item.classList.remove("checked");
        item.querySelector("input").checked = false;
        item.style.display = "";
    });
    document.getElementById("search-" + key).value = "";
    updateFace(key);
}

function updateFace(key) {
    const face        = document.getElementById("face-" + key);
    const placeholder = document.querySelector(`[data-key="${key}"]`).dataset.placeholder;
    const sel         = selections[key];
    face.textContent  = sel.size === 0 ? placeholder : sel.size === 1 ? [...sel][0] : sel.size + " selected";
}

document.addEventListener("click", () => {
    document.querySelectorAll(".cs-panel.open").forEach(p => {
        p.classList.remove("open");
        document.getElementById("face-" + p.id.replace("panel-","")).classList.remove("open");
    });
});

/* ====================================================
   API HELPERS
   ==================================================== */
function buildParams() {
    const p = new URLSearchParams();
    const cy = document.getElementById("cy").value;
    if (cy) p.append("cy", cy);
    ["division","service","sa","invoice","model","month"].forEach(key => {
        selections[key].forEach(v => p.append(key, v));
    });
    return p.toString();
}

function fmt(n) {
    return Number(n || 0).toLocaleString("en-IN", { minimumFractionDigits: 2, maximumFractionDigits: 2 });
}

function pctBadge(val) {
    if (val === null || val === undefined) return '<span class="pct flat">N/A</span>';
    const cls  = val > 0 ? "up" : val < 0 ? "down" : "flat";
    const sign = val > 0 ? "▲ +" : val < 0 ? "▼ " : "";
    return `<span class="pct ${cls}">${sign}${val.toFixed(2)}%</span>`;
}

/* ====================================================
   PAGE 1 — FILTERS, CARDS, TABLE
   ==================================================== */
async function loadFilters() {
    try {
        const data = await fetch("/filters").then(r => r.json());
        // Page 1 dropdowns
        fillCustomSelect("division", data.division);
        fillCustomSelect("service",  data.service);
        fillCustomSelect("sa",       data.sa);
        fillCustomSelect("invoice",  data.invoice);
        fillCustomSelect("month",    data.month);
        // Page 1 model
        fillCustomSelect("model",      data.model);
        // Page 2 dropdowns (same data, separate selections)
        fillCustomSelect("c-division", data.division);
        fillCustomSelect("c-service",  data.service);
        fillCustomSelect("c-sa",       data.sa);
        fillCustomSelect("c-invoice",  data.invoice);
        fillCustomSelect("c-model",    data.model);
        // Page 3 dropdowns
        fillCustomSelect("d-division", data.division);
        fillCustomSelect("d-service",  data.service);
        fillCustomSelect("d-sa",       data.sa);
        fillCustomSelect("d-invoice",  data.invoice);
        fillCustomSelect("d-model",    data.model);
    } catch(e) { console.error("Filter load failed", e); }
}

function buildCompParams() {
    const p = new URLSearchParams();
    ["c-division","c-service","c-sa","c-invoice","c-model"].forEach(key => {
        selections[key].forEach(v => p.append(key.replace("c-",""), v));
    });
    return p.toString();
}

function applyComparison() {
    compLoaded = true;
    loadComparison();
}

async function loadCards() {
    try {
        const data = await fetch("/cards?" + buildParams()).then(r => r.json());
        document.getElementById("inflow24").innerText  = Number(data.inflow24).toLocaleString("en-IN");
        document.getElementById("inflow25").innerText  = Number(data.inflow25).toLocaleString("en-IN");
        document.getElementById("labour").innerText    = data.labour;
        document.getElementById("spares").innerText    = data.spares;
        document.getElementById("total-rev").innerText = data.total;
    } catch(e) { console.error("Card load failed", e); }
}

async function loadTable() {
    const tbody = document.getElementById("tbody");
    tbody.innerHTML = '<tr><td colspan="8" class="empty-msg">Loading…</td></tr>';
    try {
        const data = await fetch("/table?" + buildParams()).then(r => r.json());
        if (!data.length) {
            tbody.innerHTML = '<tr><td colspan="8" class="empty-msg">No records found</td></tr>';
            return;
        }
        tbody.innerHTML = data.map(r => `
          <tr>
            <td>${r.Division ?? ""}</td>
            <td>${r["Repair Order#"] ?? ""}</td>
            <td>${r["SA Name"] ?? ""}</td>
            <td>${r["Vehicle Reg#"] ?? ""}</td>
            <td>${r["Service Type"] ?? ""}</td>
            <td>${fmt(r["Net Taxable Labor Amount"])}</td>
            <td>${fmt(r["Net Taxable Parts Amt"])}</td>
            <td style="font-weight:600;color:#4f6bdc;">${fmt(r["Total Revenue"])}</td>
          </tr>`).join("");
    } catch(e) {
        tbody.innerHTML = '<tr><td colspan="7" class="empty-msg">Error loading data</td></tr>';
        console.error("Table load failed", e);
    }
}

function applyFilters() { loadCards(); loadTable(); }

async function exportExcel() {
    const btn = document.querySelector(".export-btn");
    btn.disabled = true; btn.textContent = "⏳ Exporting…";
    try {
        const res  = await fetch("/export?" + buildParams());
        if (!res.ok) throw new Error("Export failed");
        const blob = await res.blob();
        const a = document.createElement("a");
        a.href = URL.createObjectURL(blob);
        a.download = "Renault_Invoice_Export.xlsx";
        document.body.appendChild(a); a.click();
        document.body.removeChild(a); URL.revokeObjectURL(a.href);
    } catch(e) { alert("Export failed: " + e.message); }
    finally { btn.disabled = false; btn.textContent = "⬇ Export Excel"; }
}

/* ====================================================
   PAGE 2 — CY COMPARISON
   ==================================================== */
async function loadComparison() {
    const wrap = document.getElementById("comp-content");
    wrap.innerHTML = '<div class="comp-loading">Loading comparison data…</div>';
    try {
        const rows = await fetch("/comparison?" + buildCompParams()).then(r => r.json());
        if (!rows.length) {
            wrap.innerHTML = '<div class="comp-loading">No data available</div>'; return;
        }

        // Totals
        const tot = { i24:0, i25:0, l24:0, l25:0, s24:0, s25:0, t24:0, t25:0 };
        rows.forEach(r => {
            tot.i24 += r.inflow24; tot.i25 += r.inflow25;
            tot.l24 += r.labour24; tot.l25 += r.labour25;
            tot.s24 += r.spares24; tot.s25 += r.spares25;
            tot.t24 += r.total24;  tot.t25 += r.total25;
        });
        const totIPct = tot.i24 ? ((tot.i25-tot.i24)/tot.i24*100).toFixed(2) : null;
        const totLPct = tot.l24 ? ((tot.l25-tot.l24)/tot.l24*100).toFixed(2) : null;
        const totSPct = tot.s24 ? ((tot.s25-tot.s24)/tot.s24*100).toFixed(2) : null;
        const totTPct = tot.t24 ? ((tot.t25-tot.t24)/tot.t24*100).toFixed(2) : null;

        function buildTable(title, cy24key, cy25key, pctKey, isCurrency) {
            const f = v => isCurrency ? "₹ " + fmt(v) : Number(v).toLocaleString("en-IN");
            const rows_html = rows.map(r => `
              <tr>
                <td>${r.month}</td>
                <td>${f(r[cy24key])}</td>
                <td>${f(r[cy25key])}</td>
                <td>${pctBadge(r[pctKey])}</td>
              </tr>`).join("");

            const t24 = isCurrency ? tot.l24 : tot.i24;
            const t25 = isCurrency ? tot.l25 : tot.i25;
            // pick right totals
            let tot24, tot25, totPct;
            if (cy24key === "inflow24")  { tot24=tot.i24; tot25=tot.i25; totPct=totIPct; }
            else if (cy24key === "labour24") { tot24=tot.l24; tot25=tot.l25; totPct=totLPct; }
            else if (cy24key === "spares24") { tot24=tot.s24; tot25=tot.s25; totPct=totSPct; }
            else { tot24=tot.t24; tot25=tot.t25; totPct=totTPct; }

            return `
              <div class="comp-section">
                <h3>${title}</h3>
                <div class="table-wrap">
                  <table class="comp-table">
                    <thead>
                      <tr>
                        <th>Month</th>
                        <th>CY 2024</th>
                        <th>CY 2025</th>
                        <th>Difference %</th>
                      </tr>
                    </thead>
                    <tbody>${rows_html}</tbody>
                    <tfoot>
                      <tr>
                        <td>Total</td>
                        <td>${f(tot24)}</td>
                        <td>${f(tot25)}</td>
                        <td>${pctBadge(parseFloat(totPct))}</td>
                      </tr>
                    </tfoot>
                  </table>
                </div>
              </div>`;
        }

        wrap.innerHTML =
            buildTable("📥 Inflow Comparison (Unique ROs)", "inflow24", "inflow25", "inflow_pct", false) +
            buildTable("🔧 Labour Revenue Comparison",      "labour24", "labour25", "labour_pct", true)  +
            buildTable("🔩 Spares Revenue Comparison",      "spares24", "spares25", "spares_pct", true)  +
            buildTable("💰 Total Revenue Comparison",       "total24",  "total25",  "total_pct",  true);

    } catch(e) {
        wrap.innerHTML = '<div class="comp-loading">Error loading comparison data</div>';
        console.error("Comparison load failed", e);
    }
}

/* ====================================================
   PAGE 3 — DIVISION-MONTH PIVOT
   ==================================================== */
function buildDivMonParams() {
    const p = new URLSearchParams();
    ["d-division","d-service","d-sa","d-invoice","d-model"].forEach(key => {
        selections[key].forEach(v => p.append(key.replace("d-",""), v));
    });
    return p.toString();
}

function applyDivMonth() { divMonLoaded = true; loadDivMonth(); }

async function loadDivMonth() {
    const wrap = document.getElementById("divmonth-content");
    wrap.innerHTML = '<div class="comp-loading">Loading…</div>';
    try {
        const { divisions, months, rows } = await fetch("/division-month?" + buildDivMonParams()).then(r => r.json());

        if (!divisions.length) {
            wrap.innerHTML = '<div class="comp-loading">No data available</div>'; return;
        }

        // Index rows by division
        const byDiv = {};
        rows.forEach(r => byDiv[r.division] = r);

        function gPct(v24, v25) {
            if (!v24) return '<span class="pct flat">N/A</span>';
            const p = ((v25 - v24) / v24 * 100);
            const cls  = p > 0 ? "up" : p < 0 ? "down" : "flat";
            const sign = p > 0 ? "▲ +" : p < 0 ? "▼ " : "";
            return `<span class="pct ${cls}">${sign}${p.toFixed(1)}%</span>`;
        }

        function buildPivot(title, prefix, isCurrency) {
            const f = v => isCurrency
                ? fmt(v)
                : Number(v||0).toLocaleString("en-IN");

            // Header row 1: Division | Jan(span3) | Feb(span3) | … | Total(span3)
            let hdr1 = `<th rowspan="2" style="min-width:80px;">Division</th>`;
            months.forEach(m => hdr1 += `<th colspan="3">${m}</th>`);
            hdr1 += `<th colspan="3">Total</th>`;

            // Header row 2: CY24 | CY25 | Growth% per month + total
            let hdr2 = "";
            [...months, "Total"].forEach(() =>
                hdr2 += `<th>CY24</th><th>CY25</th><th>Growth%</th>`);

            // Data rows
            let body = "";
            const gTot24 = new Array(months.length).fill(0);
            const gTot25 = new Array(months.length).fill(0);

            divisions.forEach(div => {
                const r = byDiv[div];
                let cells = "";
                let rSum24 = 0, rSum25 = 0;
                months.forEach((m, mi) => {
                    const v24 = r ? (r[`${prefix}24_${m}`] || 0) : 0;
                    const v25 = r ? (r[`${prefix}25_${m}`] || 0) : 0;
                    rSum24 += v24; rSum25 += v25;
                    gTot24[mi] += v24; gTot25[mi] += v25;
                    cells += `<td>${f(v24)}</td><td>${f(v25)}</td><td>${gPct(v24,v25)}</td>`;
                });
                body += `<tr>
                    <td>${div}</td>
                    ${cells}
                    <td style="font-weight:700;">${f(rSum24)}</td>
                    <td style="font-weight:700;color:#4f6bdc;">${f(rSum25)}</td>
                    <td>${gPct(rSum24, rSum25)}</td>
                  </tr>`;
            });

            // Footer: Grand Total
            let footCells = "";
            let gSum24 = 0, gSum25 = 0;
            gTot24.forEach((v, i) => {
                gSum24 += v; gSum25 += gTot25[i];
                footCells += `<td>${f(v)}</td><td>${f(gTot25[i])}</td><td>${gPct(v, gTot25[i])}</td>`;
            });

            return `
              <div class="pivot-section comp-section">
                <h3>${title}</h3>
                <div class="pivot-wrap">
                  <table class="pivot-table">
                    <thead>
                      <tr>${hdr1}</tr>
                      <tr>${hdr2}</tr>
                    </thead>
                    <tbody>${body}</tbody>
                    <tfoot>
                      <tr>
                        <td>Grand Total</td>
                        ${footCells}
                        <td>${f(gSum24)}</td>
                        <td>${f(gSum25)}</td>
                        <td>${gPct(gSum24, gSum25)}</td>
                      </tr>
                    </tfoot>
                  </table>
                </div>
              </div>`;
        }

        wrap.innerHTML =
            buildPivot("📥 Inflow — Division × Month (CY24 vs CY25)",         "i", false) +
            buildPivot("🔧 Labour — Division × Month (CY24 vs CY25)",         "l", true)  +
            buildPivot("🔩 Spares — Division × Month (CY24 vs CY25)",         "s", true)  +
            buildPivot("💰 Total Revenue — Division × Month (CY24 vs CY25)",  "t", true);

    } catch(e) {
        wrap.innerHTML = '<div class="comp-loading">Error loading data</div>';
        console.error("DivMonth load failed", e);
    }
}
window.onload = function () {
    // Init all custom selects — Page 1 and Page 2
    document.querySelectorAll(".custom-select").forEach(buildCustomSelect);
    loadFilters();
    loadCards();
    loadTable();
};
</script>
</body>
</html>""")
# ---------------- AUTO OPEN ----------------

def open_browser():
    webbrowser.open("http://127.0.0.1:8000")


if __name__ == "__main__":
    threading.Timer(1, open_browser).start()
    uvicorn.run(app, host="127.0.0.1", port=8000)
