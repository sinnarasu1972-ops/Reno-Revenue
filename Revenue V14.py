from fastapi import FastAPI, Query
from fastapi.responses import HTMLResponse, StreamingResponse
import pandas as pd
import uvicorn
import threading
import webbrowser
import io

app = FastAPI()

# ---------------- HELPERS ----------------

def indian_format(n: float) -> str:
    """Format number in Indian style: 1,23,45,678 (no decimals)"""
    n = round(float(n))
    negative = n < 0
    n = abs(int(n))

    s = str(n)
    if len(s) <= 3:
        result = s
    else:
        result = s[-3:]
        s = s[:-3]
        while s:
            result = s[-2:] + "," + result
            s = s[:-2]

    return ("-" if negative else "") + result

CY24_PATH = r"E:\Renault\Invoice Report CY24.csv"
CY25_PATH = r"E:\Renault\Invoice Report CY25.csv"
CY26_PATH = r"E:\Renault\Invoice Report CY26.csv"  # Jan & Feb only

BRANCH_CODE_TO_NAME = {
    "AKJA": "AKOLA",
    "AUJA": "AURANGABAD",
    "AUJB": "AURANGABAD BP",
    "AVJA": "AMRAVATI",
    "BAKA": "BARAMATI",
    "CNJA": "CHANDRAPUR",
    "NAJB": "WADI",
    "NAJE": "KALAMNA",
    "NSKB": "NASHIK SATPUR",
    "PUMB": "CHINCHWAD PUNE",
    "PUME": "HADAPSAR",
    "JGJB": "JALGAON",
}


# ---------------- LOAD DATA ----------------

ORIGINAL_CSV_COLS = []  # populated by load_data()

def load_data():
    df1 = pd.read_csv(CY24_PATH, low_memory=False)
    df2 = pd.read_csv(CY25_PATH, low_memory=False)
    df3 = pd.read_csv(CY26_PATH, low_memory=False)

    df1["CY"] = "CY24"
    df2["CY"] = "CY25"
    df3["CY"] = "CY26"

    df = pd.concat([df1, df2, df3], ignore_index=True)
    df.columns = df.columns.str.strip()

    # ── Capture original CSV column order BEFORE adding computed cols ──
    global ORIGINAL_CSV_COLS
    ORIGINAL_CSV_COLS = list(df.columns)

    df["Division"] = df["Repair Order#"].astype(str).str[2:6].map(BRANCH_CODE_TO_NAME).fillna(df["Repair Order#"].astype(str).str[2:6])
    df["SA Name"]  = df["RO Owner First Name"].astype(str) + " " + df["RO Owner Last Name"].astype(str)
    df["Model"]    = df["Vehicle Model"].astype(str).str.strip().str.split().str[0].str.upper()

    # Format in CSV: "01-01-2025 10:06:00"  (DD-MM-YYYY HH:MM:SS)
    # Slice first 10 chars -> "01-01-2025", ignore time part
    for date_col in ["Invoice Date", "Repair Order Date"]:
        raw_date = df[date_col].astype(str).str.strip().str[:10]
        df[date_col] = pd.to_datetime(raw_date, format="%d-%m-%Y", errors="coerce")
        mask = df[date_col].isna()
        if mask.any():
            df.loc[mask, date_col] = pd.to_datetime(
                raw_date[mask], format="%d/%m/%Y", errors="coerce"
            )

    df["Month"] = df["Invoice Date"].dt.strftime("%b")  # Jan, Feb, Mar ...

    # TAT = Invoice Date - Repair Order Date (in days)
    df["TAT (Days)"] = (df["Invoice Date"] - df["Repair Order Date"]).dt.days

    for col in ["Net Taxable Labor Amount", "Net Taxable Parts Amt"]:
        df[col] = (
            df[col].astype(str)
            .str.replace("Rs.", "", regex=False)
            .str.replace(",",   "", regex=False)
            .pipe(pd.to_numeric, errors="coerce")
            .fillna(0)
        )

    # Split Cancelled vs Active — works whether column is "Status" or "Invoice Status"
    status_col = None
    for c in df.columns:
        if c.strip().lower() in ("status", "invoice status", "ro status"):
            status_col = c
            break

    if status_col:
        cancelled_mask = df[status_col].astype(str).str.strip().str.lower() == "cancelled"
    else:
        cancelled_mask = pd.Series(False, index=df.index)

    df_cancelled = df[cancelled_mask].copy()
    df_active    = df[~cancelled_mask].copy()

    return df_active, df_cancelled


df, df_cancelled = load_data()


# ---------------- FILTERS ----------------

@app.get("/filters")
def filters():
    return {
        "division": sorted(df["Division"].dropna().unique().tolist()),
        "service":  sorted(df["Service Type"].dropna().unique().tolist()),
        "sa":       sorted(df["SA Name"].dropna().unique().tolist()),
        "invoice":  sorted(df["Invoice Type"].dropna().unique().tolist()),
        "model":    sorted(df["Model"].dropna().unique().tolist()),
        "month":    sorted(
            df["Month"].dropna().unique().tolist(),
            key=lambda m: ["Jan","Feb","Mar","Apr","May","Jun",
                           "Jul","Aug","Sep","Oct","Nov","Dec"].index(m)
            if m in ["Jan","Feb","Mar","Apr","May","Jun",
                     "Jul","Aug","Sep","Oct","Nov","Dec"] else 99
        ),
    }



# ---------------- DEPENDENT FILTERS ----------------

@app.get("/filters-dep")
def filters_dep(
    division: list[str] = Query(None),
    cy:       str = None,
):
    """Returns filtered options based on selected divisions (and optionally CY)."""
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    data = df.copy()
    if cy:       data = data[data["CY"] == cy]
    if division: data = data[data["Division"].isin(division)]

    return {
        "service":  sorted(data["Service Type"].dropna().unique().tolist()),
        "sa":       sorted(data["SA Name"].dropna().unique().tolist()),
        "invoice":  sorted(data["Invoice Type"].dropna().unique().tolist()),
        "model":    sorted(data["Model"].dropna().unique().tolist()),
        "month":    sorted(
            data["Month"].dropna().unique().tolist(),
            key=lambda m: MONTHS.index(m) if m in MONTHS else 99
        ),
    }


# ---------------- CARDS ----------------

@app.get("/cards")
def cards(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    month:    list[str] = Query(None),
    model:    list[str] = Query(None),
    cy: str = None,
):
    data = df.copy()

    if cy:       data = data[data["CY"] == cy]
    if division: data = data[data["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]
    if month:    data = data[data["Month"].isin(month)]
    if model:    data = data[data["Model"].isin(model)]

    return {
        "inflow24": int(data[data["CY"] == "CY24"]["Repair Order#"].nunique()),
        "inflow25": int(data[data["CY"] == "CY25"]["Repair Order#"].nunique()),
        "inflow26": int(data[data["CY"] == "CY26"]["Repair Order#"].nunique()),
        "labour":   "₹ " + indian_format(data['Net Taxable Labor Amount'].sum()),
        "spares":   "₹ " + indian_format(data['Net Taxable Parts Amt'].sum()),
        "total":    "₹ " + indian_format(data['Net Taxable Labor Amount'].sum() + data['Net Taxable Parts Amt'].sum()),
    }


# ---------------- TABLE ----------------

@app.get("/table")
def table(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    month:    list[str] = Query(None),
    model:    list[str] = Query(None),
    cy: str = None,
):
    data = df.copy()

    if cy:       data = data[data["CY"] == cy]
    if division: data = data[data["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]
    if month:    data = data[data["Month"].isin(month)]
    if model:    data = data[data["Model"].isin(model)]

    keep = ["Division", "Repair Order#", "SA Name", "Vehicle Reg#",
            "Service Type", "Net Taxable Labor Amount", "Net Taxable Parts Amt"]
    out = data[keep].head(500).copy()
    out["Total Revenue"] = out["Net Taxable Labor Amount"] + out["Net Taxable Parts Amt"]
    return out.to_dict(orient="records")


# ---------------- SHARED EXPORT HELPER ----------------

def _style_sheet(ws, highlight_last_n=3):
    """Auto-fit columns and highlight last N header cells in green."""
    from openpyxl.styles import PatternFill, Font
    green_fill  = PatternFill("solid", fgColor="17A34A")
    red_fill    = PatternFill("solid", fgColor="DC2626")
    white_font  = Font(color="FFFFFF", bold=True)
    blue_fill   = PatternFill("solid", fgColor="1D4ED8")

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    total_cols = ws.max_column
    # Header row — blue for all, green for last N (data cols)
    for ci in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=ci)
        fill = green_fill if ci > total_cols - highlight_last_n else blue_fill
        cell.fill = fill
        cell.font = white_font


def _style_cancelled_sheet(ws):
    """Style the Cancelled sheet header in red."""
    from openpyxl.styles import PatternFill, Font
    red_fill   = PatternFill("solid", fgColor="DC2626")
    white_font = Font(color="FFFFFF", bold=True)
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)
    for ci in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = red_fill
        cell.font = white_font


def _build_export_cols(data):
    """Return columns in original CSV file order, with Labour/Parts/Total/TAT at the end.

    Final column layout:
      1. All original CSV columns in file order (Net Taxable Labour/Parts excluded — moved to end)
      2. Computed helper cols not in original CSV: Division, SA Name, Model, Month, CY
      3. Net Taxable Labor Amount
      4. Net Taxable Parts Amt
      5. Total  (Labour + Parts)
      6. TAT (Days)  (Invoice Date - Repair Order Date)
    Returns: (export_cols list, data with Total column added)
    """
    tail_fixed = ["Net Taxable Labor Amount", "Net Taxable Parts Amt"]
    added_cols = ["Division", "SA Name", "Model", "Month", "CY"]

    # Original CSV cols in file order, skip the two that move to tail
    csv_cols   = [c for c in ORIGINAL_CSV_COLS if c in data.columns and c not in tail_fixed]
    # Our computed helper cols not already in csv_cols
    extra_cols = [c for c in added_cols if c in data.columns and c not in csv_cols]

    export_cols = csv_cols + extra_cols

    # Append Labour, Parts
    for col in tail_fixed:
        if col in data.columns:
            export_cols.append(col)

    # Add Total (Labour + Parts) column
    data = data.copy()
    if "Net Taxable Labor Amount" in data.columns and "Net Taxable Parts Amt" in data.columns:
        data["Total"] = data["Net Taxable Labor Amount"] + data["Net Taxable Parts Amt"]
        export_cols.append("Total")

    # TAT at the very end
    if "TAT (Days)" in data.columns:
        export_cols.append("TAT (Days)")

    return export_cols, data


def _apply_standard_filters(data, canc, cy, division, service, sa, invoice, month, model):
    if cy:       data = data[data["CY"] == cy];       canc = canc[canc["CY"] == cy]
    if division: data = data[data["Division"].isin(division)]; canc = canc[canc["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]; canc = canc[canc["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)];         canc = canc[canc["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]; canc = canc[canc["Invoice Type"].isin(invoice)]
    if month:    data = data[data["Month"].isin(month)];        canc = canc[canc["Month"].isin(month)]
    if model:    data = data[data["Model"].isin(model)];        canc = canc[canc["Model"].isin(model)]
    return data, canc


def _write_cancelled_sheet(writer, canc):
    """Write Cancelled rows to a red-styled sheet in the workbook."""
    if canc.empty:
        canc_placeholder = pd.DataFrame([{"Note": "No Cancelled records for current filters."}])
        canc_placeholder.to_excel(writer, index=False, sheet_name="Cancelled")
    else:
        canc_cols, canc = _build_export_cols(canc)
        canc[canc_cols].to_excel(writer, index=False, sheet_name="Cancelled")
    _style_cancelled_sheet(writer.sheets["Cancelled"])


def _write_invoice_data_sheet(writer, data):
    """Write a full raw Invoice Data sheet (same as Page 1 export) — blue header, last 4 cols green."""
    export_cols, data = _build_export_cols(data)
    data[export_cols].to_excel(writer, index=False, sheet_name="Invoice Data")
    _style_sheet(writer.sheets["Invoice Data"], highlight_last_n=4)


def _stream_excel(output, filename):
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ---------------- EXPORT — PAGE 1 (Invoice Data) ----------------


@app.get("/export")
def export(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    month:    list[str] = Query(None),
    model:    list[str] = Query(None),
    cy: str = None,
):
    data = df.copy()
    canc = df_cancelled.copy()
    data, canc = _apply_standard_filters(data, canc, cy, division, service, sa, invoice, month, model)

    export_cols, data = _build_export_cols(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        data[export_cols].to_excel(writer, index=False, sheet_name="Invoice Data")
        _style_sheet(writer.sheets["Invoice Data"], highlight_last_n=4)
        _write_cancelled_sheet(writer, canc)

    return _stream_excel(output, "Renault_Invoice_Export.xlsx")


# ---------------- EXPORT — PAGE 2 (CY Comparison) ----------------

@app.get("/export-comparison")
def export_comparison(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    model:    list[str] = Query(None),
):
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    def _f(data):
        if division: data = data[data["Division"].isin(division)]
        if service:  data = data[data["Service Type"].isin(service)]
        if sa:       data = data[data["SA Name"].isin(sa)]
        if invoice:  data = data[data["Invoice Type"].isin(invoice)]
        if model:    data = data[data["Model"].isin(model)]
        return data

    base24 = _f(df[df["CY"] == "CY24"])
    base25 = _f(df[df["CY"] == "CY25"])
    canc   = _f(df_cancelled.copy())

    rows = []
    for month in MONTHS:
        cy24 = base24[base24["Month"] == month]
        cy25 = base25[base25["Month"] == month]
        if cy24["Repair Order#"].nunique() == 0 and cy25["Repair Order#"].nunique() == 0:
            continue
        def pct(n, o): return round((n - o) / o * 100, 2) if o else None
        rows.append({
            "Month":          month,
            "Inflow CY24":    int(cy24["Repair Order#"].nunique()),
            "Inflow CY25":    int(cy25["Repair Order#"].nunique()),
            "Inflow Growth%": pct(cy25["Repair Order#"].nunique(), cy24["Repair Order#"].nunique()),
            "Labour CY24":    round(float(cy24["Net Taxable Labor Amount"].sum()), 2),
            "Labour CY25":    round(float(cy25["Net Taxable Labor Amount"].sum()), 2),
            "Labour Growth%": pct(cy25["Net Taxable Labor Amount"].sum(), cy24["Net Taxable Labor Amount"].sum()),
            "Spares CY24":    round(float(cy24["Net Taxable Parts Amt"].sum()), 2),
            "Spares CY25":    round(float(cy25["Net Taxable Parts Amt"].sum()), 2),
            "Spares Growth%": pct(cy25["Net Taxable Parts Amt"].sum(), cy24["Net Taxable Parts Amt"].sum()),
            "Total CY24":     round(float(cy24["Net Taxable Labor Amount"].sum() + cy24["Net Taxable Parts Amt"].sum()), 2),
            "Total CY25":     round(float(cy25["Net Taxable Labor Amount"].sum() + cy25["Net Taxable Parts Amt"].sum()), 2),
            "Total Growth%":  pct(cy25["Net Taxable Labor Amount"].sum() + cy25["Net Taxable Parts Amt"].sum(),
                                  cy24["Net Taxable Labor Amount"].sum() + cy24["Net Taxable Parts Amt"].sum()),
        })

    # Build full active data for Invoice Data sheet
    inv_data = df.copy()
    inv_canc = df_cancelled.copy()
    if division: inv_data = inv_data[inv_data["Division"].isin(division)]; inv_canc = inv_canc[inv_canc["Division"].isin(division)]
    if service:  inv_data = inv_data[inv_data["Service Type"].isin(service)]; inv_canc = inv_canc[inv_canc["Service Type"].isin(service)]
    if sa:       inv_data = inv_data[inv_data["SA Name"].isin(sa)]; inv_canc = inv_canc[inv_canc["SA Name"].isin(sa)]
    if invoice:  inv_data = inv_data[inv_data["Invoice Type"].isin(invoice)]; inv_canc = inv_canc[inv_canc["Invoice Type"].isin(invoice)]
    if model:    inv_data = inv_data[inv_data["Model"].isin(model)]; inv_canc = inv_canc[inv_canc["Model"].isin(model)]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="CY Comparison")
        _style_sheet(writer.sheets["CY Comparison"], highlight_last_n=0)
        _write_invoice_data_sheet(writer, inv_data)
        _write_cancelled_sheet(writer, inv_canc)

    return _stream_excel(output, "Renault_CY_Comparison_Export.xlsx")


# ---------------- EXPORT — PAGE 3 (Division-Month CY24 vs CY25) ----------------

@app.get("/export-division-month")
def export_division_month(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    model:    list[str] = Query(None),
):
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    data = df.copy()
    canc = df_cancelled.copy()
    if division: data = data[data["Division"].isin(division)]; canc = canc[canc["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]; canc = canc[canc["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)];         canc = canc[canc["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]; canc = canc[canc["Invoice Type"].isin(invoice)]
    if model:    data = data[data["Model"].isin(model)];        canc = canc[canc["Model"].isin(model)]

    divisions = sorted(data["Division"].dropna().unique().tolist())
    rows = []
    for div in divisions:
        dd = data[data["Division"] == div]
        row = {"Division": div}
        for m in MONTHS:
            m24 = dd[(dd["CY"] == "CY24") & (dd["Month"] == m)]
            m25 = dd[(dd["CY"] == "CY25") & (dd["Month"] == m)]
            row[f"Inflow CY24 {m}"] = int(m24["Repair Order#"].nunique())
            row[f"Inflow CY25 {m}"] = int(m25["Repair Order#"].nunique())
            row[f"Labour CY24 {m}"] = round(float(m24["Net Taxable Labor Amount"].sum()), 2)
            row[f"Labour CY25 {m}"] = round(float(m25["Net Taxable Labor Amount"].sum()), 2)
            row[f"Spares CY24 {m}"] = round(float(m24["Net Taxable Parts Amt"].sum()), 2)
            row[f"Spares CY25 {m}"] = round(float(m25["Net Taxable Parts Amt"].sum()), 2)
        rows.append(row)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Division-Month")
        _style_sheet(writer.sheets["Division-Month"], highlight_last_n=0)
        _write_invoice_data_sheet(writer, data)
        _write_cancelled_sheet(writer, canc)

    return _stream_excel(output, "Renault_DivisionMonth_Export.xlsx")


# ---------------- EXPORT — PAGE 4 (One Pager) ----------------

@app.get("/export-one-pager")
def export_one_pager(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    month:    list[str] = Query(None),
    model:    list[str] = Query(None),
    cy:       str = None,
):
    data = df.copy()
    canc = df_cancelled.copy()
    data, canc = _apply_standard_filters(data, canc, cy, division, service, sa, invoice, month, model)

    def cat(st):
        s = str(st).strip().upper()
        if s.startswith("PDI"):        return "PDI"
        if "ACCESSOR" in s:            return "ACC"
        if s in ("1FS",):              return "1FS"
        if s in ("2FS",):              return "2FS"
        if s in ("3FS",):              return "3FS"
        if s in ("PS","PAID SERVICE"): return "PS"
        if s in ("RR",):               return "RR"
        if s in ("B&P","BP","BODY & PAINT","BODY AND PAINT","BODYSHOP"): return "BP"
        return "OTHER"

    data = data.copy()
    data["_cat"] = data["Service Type"].apply(cat)
    divisions = sorted(data["Division"].dropna().unique().tolist())
    rows = []
    for div in divisions + ["TOTAL"]:
        d = data if div == "TOTAL" else data[data["Division"] == div]
        pdi  = d[d["_cat"] == "PDI"];  acc = d[d["_cat"] == "ACC"]
        fs1  = d[d["_cat"] == "1FS"]; fs2 = d[d["_cat"] == "2FS"]; fs3 = d[d["_cat"] == "3FS"]
        ps   = d[d["_cat"] == "PS"];   rr  = d[d["_cat"] == "RR"];  bp  = d[d["_cat"] == "BP"]
        mech = d[d["_cat"].isin(["1FS","2FS","3FS","PS","RR"])]
        mech_ro = int(mech["Repair Order#"].nunique()); bp_ro = int(bp["Repair Order#"].nunique())
        def sd(a, b): return round(a/b, 2) if b else 0
        rows.append({
            "Division":           div,
            "PDI RO":             int(pdi["Repair Order#"].nunique()),
            "Accessories RO":     int(acc["Repair Order#"].nunique()),
            "Total Billed RO":    int(mech_ro + bp_ro + pdi["Repair Order#"].nunique() + acc["Repair Order#"].nunique()),
            "1FS RO":             int(fs1["Repair Order#"].nunique()),
            "2FS RO":             int(fs2["Repair Order#"].nunique()),
            "3FS RO":             int(fs3["Repair Order#"].nunique()),
            "PS RO":              int(ps["Repair Order#"].nunique()),
            "RR RO":              int(rr["Repair Order#"].nunique()),
            "Mechanical RO":      mech_ro,
            "B&P RO":             bp_ro,
            "Total RO":           mech_ro + bp_ro,
            "Service Labour":     round(float(mech["Net Taxable Labor Amount"].sum()), 2),
            "Bodyshop Labour":    round(float(bp["Net Taxable Labor Amount"].sum()), 2),
            "Mech Parts":         round(float(mech["Net Taxable Parts Amt"].sum()), 2),
            "BP Parts":           round(float(bp["Net Taxable Parts Amt"].sum()), 2),
            "Mech Parts/RO":      sd(float(mech["Net Taxable Parts Amt"].sum()), mech_ro),
            "Mech Labour/RO":     sd(float(mech["Net Taxable Labor Amount"].sum()), mech_ro),
            "BP Parts/RO":        sd(float(bp["Net Taxable Parts Amt"].sum()), bp_ro),
            "BP Labour/RO":       sd(float(bp["Net Taxable Labor Amount"].sum()), bp_ro),
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="One Pager")
        _style_sheet(writer.sheets["One Pager"], highlight_last_n=0)
        _write_invoice_data_sheet(writer, data)
        _write_cancelled_sheet(writer, canc)

    return _stream_excel(output, "Renault_OnePager_Export.xlsx")


# ---------------- EXPORT — PAGE 5 (Current Month) ----------------

@app.get("/export-current-month")
def export_current_month(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    model:    list[str] = Query(None),
):
    import requests as req
    try:
        resp = req.get(
            "https://docs.google.com/spreadsheets/d/e/2PACX-1vSqiJ-d8D6IFLqWoBSwYyDG5-gewEzAob_CvM6CGC-Y8u_VAe_u8YklXn5nzR3DwtJBMNaxJQCf_Zmr/pub?output=csv",
            timeout=5
        )
        resp.raise_for_status()
        from io import StringIO
        gdf = pd.read_csv(StringIO(resp.text), low_memory=False)
        gdf.columns = gdf.columns.str.strip()
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=503, content={"error": "Google Sheet unavailable: " + str(e)})

    col_map = {}
    for c in gdf.columns:
        cl = c.strip().lower()
        if "invoice date" in cl:              col_map[c] = "Invoice Date"
        elif "repair order#" in cl or "repair order #" in cl: col_map[c] = "Repair Order#"
        elif "repair order date" in cl:       col_map[c] = "Repair Order Date"
        elif "vehicle model" in cl:           col_map[c] = "Vehicle Model"
        elif "vehicle reg" in cl:             col_map[c] = "Vehicle Reg#"
        elif "service type" in cl:            col_map[c] = "Service Type"
        elif "invoice type" in cl:            col_map[c] = "Invoice Type"
        elif "ro owner first" in cl:          col_map[c] = "RO Owner First Name"
        elif "ro owner last" in cl:           col_map[c] = "RO Owner Last Name"
        elif "net taxable labor" in cl:       col_map[c] = "Net Taxable Labor Amount"
        elif "net taxable parts" in cl:       col_map[c] = "Net Taxable Parts Amt"
        elif cl in ("status", "invoice status", "ro status"): col_map[c] = "Status"
    gdf = gdf.rename(columns=col_map)

    if "Repair Order#" in gdf.columns:
        _raw_div = gdf["Repair Order#"].astype(str).str[2:6]
        gdf["Division"] = _raw_div.map(BRANCH_CODE_TO_NAME).fillna(_raw_div)
    if "RO Owner First Name" in gdf.columns and "RO Owner Last Name" in gdf.columns:
        gdf["SA Name"] = gdf["RO Owner First Name"].astype(str) + " " + gdf["RO Owner Last Name"].astype(str)
    if "Vehicle Model" in gdf.columns:
        gdf["Model"] = gdf["Vehicle Model"].astype(str).str.strip().str.split().str[0].str.upper()

    for col in ["Net Taxable Labor Amount", "Net Taxable Parts Amt"]:
        if col in gdf.columns:
            gdf[col] = pd.to_numeric(gdf[col].astype(str).str.replace("Rs.", "", regex=False).str.replace(",", "", regex=False), errors="coerce").fillna(0)
        else:
            gdf[col] = 0.0

    # Split cancelled
    if "Status" in gdf.columns:
        canc_mask = gdf["Status"].astype(str).str.strip().str.lower() == "cancelled"
    else:
        canc_mask = pd.Series(False, index=gdf.index)
    gdf_canc   = gdf[canc_mask].copy()
    gdf_active = gdf[~canc_mask].copy()

    if division and "Division" in gdf_active.columns: gdf_active = gdf_active[gdf_active["Division"].isin(division)]
    if service  and "Service Type" in gdf_active.columns: gdf_active = gdf_active[gdf_active["Service Type"].isin(service)]
    if sa       and "SA Name" in gdf_active.columns:      gdf_active = gdf_active[gdf_active["SA Name"].isin(sa)]
    if model    and "Model" in gdf_active.columns:        gdf_active = gdf_active[gdf_active["Model"].isin(model)]

    # Drop any internal helper cols before export
    gdf_active = gdf_active.drop(columns=[c for c in ["_cat","Total Revenue"] if c in gdf_active.columns])
    gdf_export_cols, gdf_active = _build_export_cols(gdf_active)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        gdf_active[gdf_export_cols].to_excel(writer, index=False, sheet_name="Invoice Data")
        _style_sheet(writer.sheets["Invoice Data"], highlight_last_n=4)
        _write_cancelled_sheet(writer, gdf_canc)

    return _stream_excel(output, "Renault_CurrentMonth_Export.xlsx")


# ---------------- EXPORT — PAGE 6 (Division-Month CY25 vs CY26) ----------------

@app.get("/export-division-month-cy26")
def export_division_month_cy26(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    model:    list[str] = Query(None),
):
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    data = df.copy()
    canc = df_cancelled.copy()
    if division: data = data[data["Division"].isin(division)]; canc = canc[canc["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]; canc = canc[canc["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)];         canc = canc[canc["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]; canc = canc[canc["Invoice Type"].isin(invoice)]
    if model:    data = data[data["Model"].isin(model)];        canc = canc[canc["Model"].isin(model)]

    cy26_months = sorted(
        data[data["CY"] == "CY26"]["Month"].dropna().unique().tolist(),
        key=lambda m: MONTHS.index(m) if m in MONTHS else 99
    )
    divisions = sorted(data["Division"].dropna().unique().tolist())
    rows = []
    for div in divisions:
        dd = data[data["Division"] == div]
        row = {"Division": div}
        for m in cy26_months:
            m25 = dd[(dd["CY"] == "CY25") & (dd["Month"] == m)]
            m26 = dd[(dd["CY"] == "CY26") & (dd["Month"] == m)]
            row[f"Inflow CY25 {m}"] = int(m25["Repair Order#"].nunique())
            row[f"Inflow CY26 {m}"] = int(m26["Repair Order#"].nunique())
            row[f"Labour CY25 {m}"] = round(float(m25["Net Taxable Labor Amount"].sum()), 2)
            row[f"Labour CY26 {m}"] = round(float(m26["Net Taxable Labor Amount"].sum()), 2)
            row[f"Spares CY25 {m}"] = round(float(m25["Net Taxable Parts Amt"].sum()), 2)
            row[f"Spares CY26 {m}"] = round(float(m26["Net Taxable Parts Amt"].sum()), 2)
        rows.append(row)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Div-Month CY26")
        _style_sheet(writer.sheets["Div-Month CY26"], highlight_last_n=0)
        _write_invoice_data_sheet(writer, data)
        _write_cancelled_sheet(writer, canc)

    return _stream_excel(output, "Renault_DivMonth_CY26_Export.xlsx")


# ---------------- COMPARISON ----------------

@app.get("/comparison")
def comparison(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    model:    list[str] = Query(None),
):
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    def apply_filters(data):
        if division: data = data[data["Division"].isin(division)]
        if service:  data = data[data["Service Type"].isin(service)]
        if sa:       data = data[data["SA Name"].isin(sa)]
        if invoice:  data = data[data["Invoice Type"].isin(invoice)]
        if model:    data = data[data["Model"].isin(model)]
        return data

    base24 = apply_filters(df[df["CY"] == "CY24"])
    base25 = apply_filters(df[df["CY"] == "CY25"])

    result = []
    for month in MONTHS:
        cy24 = base24[base24["Month"] == month]
        cy25 = base25[base25["Month"] == month]

        inflow24 = int(cy24["Repair Order#"].nunique())
        inflow25 = int(cy25["Repair Order#"].nunique())
        labour24 = float(cy24["Net Taxable Labor Amount"].sum())
        labour25 = float(cy25["Net Taxable Labor Amount"].sum())
        spares24 = float(cy24["Net Taxable Parts Amt"].sum())
        spares25 = float(cy25["Net Taxable Parts Amt"].sum())

        def pct(new, old):
            if old == 0: return None
            return round((new - old) / old * 100, 2)

        if inflow24 == 0 and inflow25 == 0:
            continue

        result.append({
            "month":      month,
            "inflow24":   inflow24,  "inflow25":   inflow25,  "inflow_pct": pct(inflow25, inflow24),
            "labour24":   labour24,  "labour25":   labour25,  "labour_pct": pct(labour25, labour24),
            "spares24":   spares24,  "spares25":   spares25,  "spares_pct": pct(spares25, spares24),
            "total24":    labour24 + spares24,
            "total25":    labour25 + spares25,
            "total_pct":  pct(labour25 + spares25, labour24 + spares24),
        })

    return result


# ---------------- DIVISION-MONTH ----------------

@app.get("/division-month")
def division_month(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    model:    list[str] = Query(None),
):
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    data = df.copy()
    if division: data = data[data["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]
    if model:    data = data[data["Model"].isin(model)]

    divisions = sorted(data["Division"].dropna().unique().tolist())

    # Build pivot: for each division, one row with values per month
    rows = []
    for div in divisions:
        div_data = data[data["Division"] == div]
        row = {"division": div}
        for month in MONTHS:
            m24 = div_data[(div_data["CY"] == "CY24") & (div_data["Month"] == month)]
            m25 = div_data[(div_data["CY"] == "CY25") & (div_data["Month"] == month)]
            row[f"i24_{month}"] = int(m24["Repair Order#"].nunique())
            row[f"i25_{month}"] = int(m25["Repair Order#"].nunique())
            row[f"l24_{month}"] = float(m24["Net Taxable Labor Amount"].sum())
            row[f"l25_{month}"] = float(m25["Net Taxable Labor Amount"].sum())
            row[f"s24_{month}"] = float(m24["Net Taxable Parts Amt"].sum())
            row[f"s25_{month}"] = float(m25["Net Taxable Parts Amt"].sum())
            row[f"t24_{month}"] = row[f"l24_{month}"] + row[f"s24_{month}"]
            row[f"t25_{month}"] = row[f"l25_{month}"] + row[f"s25_{month}"]
        rows.append(row)

    return {"divisions": divisions, "months": MONTHS, "rows": rows}



# ---------------- DIVISION-MONTH CY25 vs CY26 ----------------

@app.get("/division-month-cy26")
def division_month_cy26(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    model:    list[str] = Query(None),
):
    try:
        MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

        data = df.copy()
        if division: data = data[data["Division"].isin(division)]
        if service:  data = data[data["Service Type"].isin(service)]
        if sa:       data = data[data["SA Name"].isin(sa)]
        if invoice:  data = data[data["Invoice Type"].isin(invoice)]
        if model:    data = data[data["Model"].isin(model)]

        # Only show months that exist in CY26
        cy26_months = sorted(
            data[data["CY"] == "CY26"]["Month"].dropna().unique().tolist(),
            key=lambda m: MONTHS.index(m) if m in MONTHS else 99
        )

        divisions = sorted(data["Division"].dropna().unique().tolist())

        rows = []
        for div in divisions:
            div_data = data[data["Division"] == div]
            row = {"division": div}
            for month in cy26_months:
                m25 = div_data[(div_data["CY"] == "CY25") & (div_data["Month"] == month)]
                m26 = div_data[(div_data["CY"] == "CY26") & (div_data["Month"] == month)]
                row[f"i25_{month}"] = int(m25["Repair Order#"].nunique())
                row[f"i26_{month}"] = int(m26["Repair Order#"].nunique())
                row[f"l25_{month}"] = float(m25["Net Taxable Labor Amount"].sum())
                row[f"l26_{month}"] = float(m26["Net Taxable Labor Amount"].sum())
                row[f"s25_{month}"] = float(m25["Net Taxable Parts Amt"].sum())
                row[f"s26_{month}"] = float(m26["Net Taxable Parts Amt"].sum())
                row[f"t25_{month}"] = row[f"l25_{month}"] + row[f"s25_{month}"]
                row[f"t26_{month}"] = row[f"l26_{month}"] + row[f"s26_{month}"]
            rows.append(row)

        return {"divisions": divisions, "months": cy26_months, "rows": rows}
    except Exception as e:
        return {"divisions": [], "months": [], "rows": [], "error": str(e)}


# ---------------- ONE PAGER REPORT ----------------

@app.get("/one-pager")
def one_pager(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    invoice:  list[str] = Query(None),
    month:    list[str] = Query(None),
    model:    list[str] = Query(None),
    cy:       str = None,
):
    # ── If CY26 requested, check which months need Google Sheet ────────────
    # CY26 CSV has Jan+Feb only. Any other month (e.g. Mar) must come from GSheet.
    cy26_csv_months = set(df[df["CY"] == "CY26"]["Month"].dropna().unique().tolist())
    need_gsheet = (
        cy == "CY26" and
        month and
        any(m not in cy26_csv_months for m in month)
    )
    # Also fetch if CY26 + no month filter (user wants all CY26 including live month)
    need_gsheet = need_gsheet or (cy == "CY26" and not month)

    gsheet_df = None
    if need_gsheet:
        try:
            import requests as req
            from io import StringIO
            resp = req.get(GSHEET_CSV, timeout=5)
            resp.raise_for_status()
            gdf = pd.read_csv(StringIO(resp.text), low_memory=False)
            gdf.columns = gdf.columns.str.strip()

            col_map = {}
            for c in gdf.columns:
                cl = c.strip().lower()
                if "invoice date" in cl:                             col_map[c] = "Invoice Date"
                elif "repair order#" in cl or "repair order #" in cl: col_map[c] = "Repair Order#"
                elif "repair order date" in cl:                      col_map[c] = "Repair Order Date"
                elif "vehicle model" in cl:                          col_map[c] = "Vehicle Model"
                elif "vehicle reg" in cl:                            col_map[c] = "Vehicle Reg#"
                elif "service type" in cl:                           col_map[c] = "Service Type"
                elif "invoice type" in cl:                           col_map[c] = "Invoice Type"
                elif "ro owner first" in cl:                         col_map[c] = "RO Owner First Name"
                elif "ro owner last" in cl:                          col_map[c] = "RO Owner Last Name"
                elif "net taxable labor" in cl:                      col_map[c] = "Net Taxable Labor Amount"
                elif "net taxable parts" in cl:                      col_map[c] = "Net Taxable Parts Amt"
            gdf = gdf.rename(columns=col_map)

            if "Repair Order#" in gdf.columns:
                _raw_div = gdf["Repair Order#"].astype(str).str[2:6]
                gdf["Division"] = _raw_div.map(BRANCH_CODE_TO_NAME).fillna(_raw_div)
            if "RO Owner First Name" in gdf.columns and "RO Owner Last Name" in gdf.columns:
                gdf["SA Name"] = gdf["RO Owner First Name"].astype(str) + " " + gdf["RO Owner Last Name"].astype(str)
            if "Vehicle Model" in gdf.columns:
                gdf["Model"] = gdf["Vehicle Model"].astype(str).str.strip().str.split().str[0].str.upper()

            # Parse invoice date for Month
            raw_date = gdf["Invoice Date"].astype(str).str.strip().str[:10]
            gdf["Invoice Date"] = pd.to_datetime(raw_date, format="%d-%m-%Y", errors="coerce")
            mask = gdf["Invoice Date"].isna()
            if mask.any():
                gdf.loc[mask, "Invoice Date"] = pd.to_datetime(raw_date[mask], format="%d/%m/%Y", errors="coerce")
            gdf["Month"] = gdf["Invoice Date"].dt.strftime("%b")
            gdf["CY"] = "CY26"

            for col in ["Net Taxable Labor Amount", "Net Taxable Parts Amt"]:
                if col in gdf.columns:
                    gdf[col] = (
                        gdf[col].astype(str)
                        .str.replace("Rs.", "", regex=False)
                        .str.replace(",", "", regex=False)
                        .pipe(pd.to_numeric, errors="coerce")
                        .fillna(0)
                    )
                else:
                    gdf[col] = 0.0

            # Only keep months NOT already in CY26 CSV (avoid duplicates)
            gsheet_df = gdf[~gdf["Month"].isin(cy26_csv_months)].copy()
        except Exception:
            gsheet_df = None  # silently fall back to CSV only

    # ── Build data ──────────────────────────────────────────────────────────
    data = df.copy()
    if gsheet_df is not None and len(gsheet_df):
        data = pd.concat([data, gsheet_df], ignore_index=True)

    if cy:       data = data[data["CY"] == cy]
    if division: data = data[data["Division"].isin(division)]
    if service:  data = data[data["Service Type"].isin(service)]
    if sa:       data = data[data["SA Name"].isin(sa)]
    if invoice:  data = data[data["Invoice Type"].isin(invoice)]
    if month:    data = data[data["Month"].isin(month)]
    if model:    data = data[data["Model"].isin(model)]

    divisions = sorted(data["Division"].dropna().unique().tolist())

    # Service Type category mapping (adjust keywords to match your actual data)
    def is_pdi(st):         return str(st).strip().upper().startswith("PDI")
    def is_acc(st):         return str(st).strip().upper().startswith("ACC") or "ACCESSORIES" in str(st).upper()
    def is_1fs(st):         return str(st).strip().upper() in ["1FS","1 FS","1FREE","1 FREE SERVICE","1FS"]
    def is_2fs(st):         return str(st).strip().upper() in ["2FS","2 FS","2FREE","2 FREE SERVICE","2FS"]
    def is_3fs(st):         return str(st).strip().upper() in ["3FS","3 FS","3FREE","3 FREE SERVICE","3FS"]
    def is_ps(st):          return str(st).strip().upper() in ["PS","PAID SERVICE","PAID SERVICES","BP"]
    def is_rr(st):          return str(st).strip().upper() in ["RR","RUNNING REPAIR","RR"]
    def is_bp(st):          return str(st).strip().upper() in ["B&P","BP","BODY & PAINT","BODY AND PAINT","BODYSHOP","BODY SHOP"]

    def cat(row):
        st = str(row.get("Service Type","")).strip().upper()
        if st.startswith("PDI"):        return "PDI"
        if "ACCESSOR" in st:            return "ACC"
        if st in ("1FS",):              return "1FS"
        if st in ("2FS",):              return "2FS"
        if st in ("3FS",):              return "3FS"
        if st in ("PS","PAID SERVICE"): return "PS"
        if st in ("RR",):               return "RR"
        if st in ("B&P","BP","BODY & PAINT","BODY AND PAINT","BODYSHOP"): return "BP"
        return "OTHER"

    data = data.copy()
    data["_cat"] = data["Service Type"].apply(lambda st: cat({"Service Type": st}))

    result = []
    ALL = "__ALL__"
    div_list = divisions + [ALL]

    for div in div_list:
        d = data if div == ALL else data[data["Division"] == div]

        pdi  = d[d["_cat"] == "PDI"]
        acc  = d[d["_cat"] == "ACC"]
        fs1  = d[d["_cat"] == "1FS"]
        fs2  = d[d["_cat"] == "2FS"]
        fs3  = d[d["_cat"] == "3FS"]
        ps   = d[d["_cat"] == "PS"]
        rr   = d[d["_cat"] == "RR"]
        bp   = d[d["_cat"] == "BP"]

        mech = d[d["_cat"].isin(["1FS","2FS","3FS","PS","RR"])]
        mech_ro  = int(mech["Repair Order#"].nunique())
        bp_ro    = int(bp["Repair Order#"].nunique())
        pdi_ro   = int(pdi["Repair Order#"].nunique())
        acc_ro   = int(acc["Repair Order#"].nunique())
        total_ro = mech_ro + bp_ro

        # Labour: separate Bodyshop labour from service labour
        # Bodyshop = BP category; Service = everything else
        service_labour = float(mech["Net Taxable Labor Amount"].sum()) +                          float(ps["Net Taxable Labor Amount"].sum()) +                          float(rr["Net Taxable Labor Amount"].sum()) +                          float(fs1["Net Taxable Labor Amount"].sum()) +                          float(fs2["Net Taxable Labor Amount"].sum()) +                          float(fs3["Net Taxable Labor Amount"].sum())
        service_labour = float(mech["Net Taxable Labor Amount"].sum())
        bodyshop_labour = float(bp["Net Taxable Labor Amount"].sum())

        mech_parts = float(mech["Net Taxable Parts Amt"].sum())
        bp_parts   = float(bp["Net Taxable Parts Amt"].sum())

        def safe_div(a, b): return round(a/b, 2) if b else 0

        result.append({
            "division":         div if div != ALL else "TOTAL",
            "pdi_ro":           pdi_ro,
            "acc_ro":           acc_ro,
            "total_billed_ro":  total_ro + pdi_ro + acc_ro,
            "fs1_ro":           int(fs1["Repair Order#"].nunique()),
            "fs2_ro":           int(fs2["Repair Order#"].nunique()),
            "fs3_ro":           int(fs3["Repair Order#"].nunique()),
            "ps_ro":            int(ps["Repair Order#"].nunique()),
            "rr_ro":            int(rr["Repair Order#"].nunique()),
            "mech_ro":          mech_ro,
            "bp_ro":            bp_ro,
            "total_ro":         total_ro,
            "service_labour":   service_labour,
            "bodyshop_labour":  bodyshop_labour,
            "mech_parts":       mech_parts,
            "bp_parts":         bp_parts,
            "mech_parts_per_ro": safe_div(mech_parts, mech_ro),
            "mech_labour_per_ro":safe_div(service_labour, mech_ro),
            "bp_parts_per_ro":   safe_div(bp_parts, bp_ro),
            "bp_labour_per_ro":  safe_div(bodyshop_labour, bp_ro),
        })

    return {"divisions": divisions, "rows": result}



# ---------------- CM.JS (Current Month JS served separately) ----------------

CM_JS = r"""
var _GSHEET_CSV = "__GSHEET_URL__";

async function _loadCurrentMonth() {
    var tbody = document.getElementById("cm-tbody");
    var msg = document.createElement("tr");
    msg.innerHTML = "<td colspan=8 class=empty-msg>Reading Google Sheet data...</td>";
    tbody.innerHTML = ""; tbody.appendChild(msg);
    document.getElementById("cm-inflow").innerText = "...";
    document.getElementById("cm-labour").innerText = "...";
    document.getElementById("cm-spares").innerText = "...";
    document.getElementById("cm-total").innerText  = "...";
    document.getElementById("cm-mtd-wrap").style.display = "none";

    // Step 1: fetch GSheet to find max invoice date
    var gsheetMaxDay = null;
    try {
        var gr = await fetch(_GSHEET_CSV);
        if (!gr.ok) throw new Error("HTTP " + gr.status);
        var csvText = await gr.text();
        var NL = String.fromCharCode(10);
        var csvRows = csvText.replace(String.fromCharCode(13), "").split(NL);
        var hdrs = csvRows[0].split(",").map(function(h) {
            return h.replace(new RegExp('"', 'g'), "").trim().toLowerCase();
        });
        var dcol = -1;
        for (var hi = 0; hi < hdrs.length; hi++) {
            if (hdrs[hi].indexOf("invoice date") !== -1) { dcol = hi; break; }
        }
        if (dcol >= 0) {
            var maxD = 0;
            for (var ri = 1; ri < csvRows.length; ri++) {
                var cell = (csvRows[ri].split(",")[dcol] || "")
                    .replace(new RegExp('"', 'g'), "").trim();
                var dd = parseInt(cell.split(" ")[0].split("-")[0], 10);
                if (!isNaN(dd) && dd > maxD && dd <= 31) maxD = dd;
            }
            if (maxD > 0) gsheetMaxDay = maxD;
        }
    } catch(ex) {
        console.warn("GSheet pre-read failed:", ex.message);
    }

    var msg2 = document.createElement("tr");
    msg2.innerHTML = "<td colspan=8 class=empty-msg>Loading MTD comparison...</td>";
    tbody.innerHTML = ""; tbody.appendChild(msg2);

    // Step 2: call backend with GSheet max day
    var params = buildCmParams();
    if (gsheetMaxDay) params += (params ? "&" : "") + "mtd_day=" + gsheetMaxDay;

    try {
        var resp = await fetch("/current-month?" + params);
        var data = await resp.json();

        if (data.error) {
            var er = document.createElement("tr");
            er.innerHTML = "<td colspan=8 class=empty-msg style=color:#ef4444>Error: " + data.error + "</td>";
            tbody.innerHTML = ""; tbody.appendChild(er); return;
        }

        var s = data.summary || {};
        document.getElementById("cm-month-label").innerText = data.month || "Current Month";
        document.getElementById("cm-inflow").innerText = Number(s.inflow||0).toLocaleString("en-IN");
        document.getElementById("cm-labour").innerText = "\u20b9 " + (s.labour||"0");
        document.getElementById("cm-spares").innerText = "\u20b9 " + (s.spares||"0");
        document.getElementById("cm-total").innerText  = "\u20b9 " + (s.total||"0");

        var now = new Date();
        var di = gsheetMaxDay ? " (GSheet up to day-" + gsheetMaxDay + ")" : "";
        document.getElementById("cm-last-updated").innerText = "Last updated: " + now.toLocaleTimeString("en-IN") + di;

        var mtd = data.mtd;
        if (mtd) {
            document.getElementById("cm-mtd-label").innerText    = mtd.label || "";
            document.getElementById("cm-mtd-cy25-hdr").innerText = "CY25 MTD (1-" + mtd.mtd_day + ")";
            document.getElementById("cm-mtd-cy26-hdr").innerText = "CY26 MTD (1-" + mtd.mtd_day + ")";

            function badge(p) {
                if (p === null || p === undefined) return "<span class=pct-flat>N/A</span>";
                var cls = p > 0 ? "pct-up" : p < 0 ? "pct-down" : "pct-flat";
                var sign = p > 0 ? "\u25b2 +" : p < 0 ? "\u25bc " : "";
                return "<span class=" + cls + ">" + sign + p.toFixed(1) + "%</span>";
            }

            var mrows = [
                ["Inflow (ROs)",   mtd.cy25_inflow, mtd.cy26_inflow, mtd.inflow_pct, false],
                ["Labour Revenue", mtd.cy25_labour, mtd.cy26_labour, mtd.labour_pct, true],
                ["Spares Revenue", mtd.cy25_spares, mtd.cy26_spares, mtd.spares_pct, true],
                ["Total Revenue",  mtd.cy25_total,  mtd.cy26_total,  mtd.total_pct,  true],
            ];
            var mh = "";
            mrows.forEach(function(r) {
                var v25 = r[4] ? ("\u20b9 " + r[1]) : Number(r[1]||0).toLocaleString("en-IN");
                var v26 = r[4] ? ("\u20b9 " + r[2]) : Number(r[2]||0).toLocaleString("en-IN");
                mh += "<tr><td style=text-align:left;font-weight:600>" + r[0] + "</td>" +
                    "<td>" + v25 + "</td><td style=color:#15803d;font-weight:700>" + v26 + "</td>" +
                    "<td>" + badge(r[3]) + "</td></tr>";
            });
            document.getElementById("cm-mtd-tbody").innerHTML = mh;
            document.getElementById("cm-mtd-wrap").style.display = "block";
        }

        var rows = data.rows || [];
        var mtdDayLimit = (data.mtd && data.mtd.mtd_day) ? data.mtd.mtd_day : (gsheetMaxDay || 31);

        // ── Division summary table (CY25 MTD vs CY26 MTD comparison) ──
        // Use backend-computed division breakdown (full GSheet, filtered to MTD day)
        var divCy25 = data.div_cy25 || {};
        var divCy26 = data.div_cy26 || {};

        // Fallback: compute from rows if backend didn't provide div_cy26
        if (Object.keys(divCy26).length === 0) {
            var mtdRows = rows.filter(function(r) {
                var d = (r["Invoice Date"] || "").split(" ")[0];
                var day = parseInt((d.split("-")[0] || "0"), 10);
                return day >= 1 && day <= mtdDayLimit;
            });
            if (mtdRows.length === 0) mtdRows = rows;
            mtdRows.forEach(function(r) {
                var div = r.Division || "Unknown";
                var ro  = r["Repair Order#"] || "";
                var l   = parseFloat(r["Net Taxable Labor Amount"]) || 0;
                var sp  = parseFloat(r["Net Taxable Parts Amt"])    || 0;
                if (!divCy26[div]) divCy26[div] = {inflow:0, labour:0, spares:0, ros:new Set()};
                if (ro) divCy26[div].ros.add(ro);
                divCy26[div].labour += l;
                divCy26[div].spares += sp;
                divCy26[div].inflow = divCy26[div].ros.size;
            });
        }

        function gpBadge(v25, v26) {
            if (!v25) return "<span class=pct-flat>N/A</span>";
            var p = ((v26-v25)/v25*100);
            var cls = p>0?"pct-up":p<0?"pct-down":"pct-flat";
            var sign = p>0?"▲ +":p<0?"▼ ":"";
            return "<span class=" + cls + ">" + sign + p.toFixed(1) + "%</span>";
        }
        function fi(v) { return Math.round(v||0).toLocaleString("en-IN"); }
        function fc(v) { return "₹ " + Math.round(v||0).toLocaleString("en-IN"); }

        // All divisions — union of CY25 and CY26
        var allDivs = new Set(Object.keys(divCy25).concat(Object.keys(divCy26)));
        var divHtml = "";
        var gI25=0,gI26=0,gL25=0,gL26=0,gS25=0,gS26=0;

        Array.from(allDivs).sort().forEach(function(div) {
            var d25 = divCy25[div] || {inflow:0,labour:0,spares:0};
            var d26 = divCy26[div] || {inflow:0,labour:0,spares:0};
            var i25=d25.inflow, i26=d26.inflow;
            var l25=d25.labour, l26=d26.labour;
            var s25=d25.spares, s26=d26.spares;
            gI25+=i25; gI26+=i26; gL25+=l25; gL26+=l26; gS25+=s25; gS26+=s26;
            divHtml += "<tr>" +
                "<td style=text-align:left;font-weight:600;color:#4f6bdc>" + div + "</td>" +
                "<td>" + fi(i25) + "</td><td style=color:#15803d;font-weight:600>" + fi(i26) + "</td><td>" + gpBadge(i25,i26) + "</td>" +
                "<td>" + fc(l25) + "</td><td style=color:#15803d;font-weight:600>" + fc(l26) + "</td><td>" + gpBadge(l25,l26) + "</td>" +
                "<td>" + fc(s25) + "</td><td style=color:#15803d;font-weight:600>" + fc(s26) + "</td><td>" + gpBadge(s25,s26) + "</td>" +
                "<td>" + fc(l25+s25) + "</td><td style=color:#15803d;font-weight:600>" + fc(l26+s26) + "</td><td>" + gpBadge(l25+s25,l26+s26) + "</td>" +
                "</tr>";
        });
        // Total row
        divHtml += "<tr style=background:#eef1ff;font-weight:700;border-top:2px solid #4f6bdc>" +
            "<td style=text-align:left>TOTAL</td>" +
            "<td>" + fi(gI25) + "</td><td style=color:#15803d>" + fi(gI26) + "</td><td>" + gpBadge(gI25,gI26) + "</td>" +
            "<td>" + fc(gL25) + "</td><td style=color:#15803d>" + fc(gL26) + "</td><td>" + gpBadge(gL25,gL26) + "</td>" +
            "<td>" + fc(gS25) + "</td><td style=color:#15803d>" + fc(gS26) + "</td><td>" + gpBadge(gS25,gS26) + "</td>" +
            "<td>" + fc(gL25+gS25) + "</td><td style=color:#15803d>" + fc(gL26+gS26) + "</td><td>" + gpBadge(gL25+gS25,gL26+gS26) + "</td>" +
            "</tr>";

        document.getElementById("cm-div-tbody").innerHTML = divHtml;
        document.getElementById("cm-div-month-label").innerText = data.mtd_label || data.month || "Current Month";
        document.getElementById("cm-div-wrap").style.display = "block";

        if (!rows.length) {
            var nr = document.createElement("tr");
            nr.innerHTML = "<td colspan=8 class=empty-msg>No records found.</td>";
            tbody.innerHTML = ""; tbody.appendChild(nr); return;
        }
        var th = "";
        rows.forEach(function(r) {
            var l = Math.round(r["Net Taxable Labor Amount"] || 0);
            var sp = Math.round(r["Net Taxable Parts Amt"]   || 0);
            th += "<tr><td>" + (r.Division||"") + "</td><td>" + (r["Repair Order#"]||"") + "</td>" +
                "<td>" + (r["SA Name"]||"") + "</td><td>" + (r["Vehicle Reg#"]||"") + "</td>" +
                "<td>" + (r["Service Type"]||"") + "</td>" +
                "<td>" + l.toLocaleString("en-IN") + "</td>" +
                "<td>" + sp.toLocaleString("en-IN") + "</td>" +
                "<td style=font-weight:600;color:#4f6bdc>" + (l+sp).toLocaleString("en-IN") + "</td></tr>";
        });
        tbody.innerHTML = th;

    } catch(e2) {
        var er2 = document.createElement("tr");
        er2.innerHTML = "<td colspan=8 class=empty-msg style=color:#ef4444>Failed: " + e2.message + "</td>";
        tbody.innerHTML = ""; tbody.appendChild(er2);
        console.error("_loadCurrentMonth failed:", e2);
    }
}
""".replace("__GSHEET_URL__", "https://docs.google.com/spreadsheets/d/e/2PACX-1vSqiJ-d8D6IFLqWoBSwYyDG5-gewEzAob_CvM6CGC-Y8u_VAe_u8YklXn5nzR3DwtJBMNaxJQCf_Zmr/pub?output=csv")

from fastapi.responses import Response as FResponse

@app.get("/cm.js")
def cm_js():
    return FResponse(content=CM_JS, media_type="application/javascript")

# ---------------- CURRENT MONTH (Google Sheet) ----------------

GSHEET_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSqiJ-d8D6IFLqWoBSwYyDG5-gewEzAob_CvM6CGC-Y8u_VAe_u8YklXn5nzR3DwtJBMNaxJQCf_Zmr/pubhtml"
GSHEET_CSV = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSqiJ-d8D6IFLqWoBSwYyDG5-gewEzAob_CvM6CGC-Y8u_VAe_u8YklXn5nzR3DwtJBMNaxJQCf_Zmr/pub?output=csv"

@app.get("/current-month")
def current_month(
    division: list[str] = Query(None),
    service:  list[str] = Query(None),
    sa:       list[str] = Query(None),
    model:    list[str] = Query(None),
    mtd_day:  int = None,   # passed by browser after reading GSheet max date
):
    try:
        import requests as req
        resp = req.get(GSHEET_CSV, timeout=5)
        resp.raise_for_status()
        from io import StringIO
        raw = StringIO(resp.text)
        gdf = pd.read_csv(raw, low_memory=False)
        gdf.columns = gdf.columns.str.strip()
    except Exception as e:
        return {"error": "Google Sheet unavailable: " + str(e), "rows": [], "summary": {}, "mtd": None, "div_cy25": {}, "div_cy26": {}, "month": "Current Month"}

    # Normalise columns to match existing CSV schema
    col_map = {}
    for c in gdf.columns:
        cl = c.strip().lower()
        if "invoice date" in cl:              col_map[c] = "Invoice Date"
        elif "repair order#" in cl or "repair order #" in cl: col_map[c] = "Repair Order#"
        elif "repair order date" in cl:       col_map[c] = "Repair Order Date"
        elif "vehicle model" in cl:           col_map[c] = "Vehicle Model"
        elif "vehicle reg" in cl:             col_map[c] = "Vehicle Reg#"
        elif "service type" in cl:            col_map[c] = "Service Type"
        elif "invoice type" in cl:            col_map[c] = "Invoice Type"
        elif "ro owner first" in cl:          col_map[c] = "RO Owner First Name"
        elif "ro owner last" in cl:           col_map[c] = "RO Owner Last Name"
        elif "net taxable labor" in cl:       col_map[c] = "Net Taxable Labor Amount"
        elif "net taxable parts" in cl:       col_map[c] = "Net Taxable Parts Amt"
    gdf = gdf.rename(columns=col_map)

    # Derived columns
    if "Repair Order#" in gdf.columns:
        _raw_div = gdf["Repair Order#"].astype(str).str[2:6]
        gdf["Division"] = _raw_div.map(BRANCH_CODE_TO_NAME).fillna(_raw_div)
    if "RO Owner First Name" in gdf.columns and "RO Owner Last Name" in gdf.columns:
        gdf["SA Name"] = gdf["RO Owner First Name"].astype(str) + " " + gdf["RO Owner Last Name"].astype(str)
    if "Vehicle Model" in gdf.columns:
        gdf["Model"] = gdf["Vehicle Model"].astype(str).str.strip().str.split().str[0].str.upper()

    for col in ["Net Taxable Labor Amount", "Net Taxable Parts Amt"]:
        if col in gdf.columns:
            gdf[col] = (
                gdf[col].astype(str)
                .str.replace("Rs.", "", regex=False)
                .str.replace(",", "", regex=False)
                .pipe(pd.to_numeric, errors="coerce")
                .fillna(0)
            )
        else:
            gdf[col] = 0.0

    # Apply filters
    if division and "Division" in gdf.columns:
        gdf = gdf[gdf["Division"].isin(division)]
    if service and "Service Type" in gdf.columns:
        gdf = gdf[gdf["Service Type"].isin(service)]
    if sa and "SA Name" in gdf.columns:
        gdf = gdf[gdf["SA Name"].isin(sa)]
    if model and "Model" in gdf.columns:
        gdf = gdf[gdf["Model"].isin(model)]

    labour = float(gdf["Net Taxable Labor Amount"].sum())
    spares = float(gdf["Net Taxable Parts Amt"].sum())

    keep_cols = [c for c in ["Division","Repair Order#","SA Name","Vehicle Reg#",
                              "Service Type","Net Taxable Labor Amount",
                              "Net Taxable Parts Amt"] if c in gdf.columns]
    rows_out = gdf[keep_cols].head(500).copy()
    rows_out["Total Revenue"] = rows_out.get("Net Taxable Labor Amount", 0) + rows_out.get("Net Taxable Parts Amt", 0)

    # ---- MTD comparison: use GSheet latest date supplied by browser ----
    from datetime import datetime
    today = datetime.now()
    curr_month_name = today.strftime("%b")       # e.g. "Mar"
    curr_year       = today.year                 # 2026
    # mtd_day supplied by browser (max date in GSheet) — fall back to today if missing
    if not mtd_day:
        mtd_day = today.day

    # CY26 current month rows (already in gdf if filter applied)
    # For MTD we need unfiltered df too (compare same day range in CY25)
    base_df = df.copy()
    if division and "Division" in base_df.columns: base_df = base_df[base_df["Division"].isin(division)]
    if service  and "Service Type" in base_df.columns: base_df = base_df[base_df["Service Type"].isin(service)]
    if sa       and "SA Name" in base_df.columns:      base_df = base_df[base_df["SA Name"].isin(sa)]
    if model    and "Model" in base_df.columns:        base_df = base_df[base_df["Model"].isin(model)]

    cy25_mtd = base_df[
        (base_df["CY"] == "CY25") &
        (base_df["Month"] == curr_month_name) &
        (base_df["Invoice Date"].dt.day <= mtd_day)
    ]
    cy25_mtd_labour = float(cy25_mtd["Net Taxable Labor Amount"].sum())
    cy25_mtd_spares = float(cy25_mtd["Net Taxable Parts Amt"].sum())
    cy25_mtd_inflow = int(cy25_mtd["Repair Order#"].nunique()) if "Repair Order#" in cy25_mtd.columns else 0

    def pct(new, old): return round((new-old)/old*100, 1) if old else None

    # CY26 MTD division breakdown — from full GSheet (not capped at 500 rows)
    # Parse invoice date day for GSheet rows
    if "Invoice Date" in gdf.columns:
        raw_d = gdf["Invoice Date"].astype(str).str.strip().str[:10]
        gdf["_day"] = pd.to_numeric(raw_d.str.split("-").str[0], errors="coerce").fillna(0).astype(int)
    else:
        gdf["_day"] = 0

    gdf_mtd = gdf[gdf["_day"] <= mtd_day] if mtd_day else gdf

    div_cy26 = {}
    if "Division" in gdf_mtd.columns:
        for div, grp in gdf_mtd.groupby("Division"):
            div_cy26[div] = {
                "inflow":  int(grp["Repair Order#"].nunique()) if "Repair Order#" in grp.columns else 0,
                "labour":  float(grp["Net Taxable Labor Amount"].sum()),
                "spares":  float(grp["Net Taxable Parts Amt"].sum()),
            }

    # Division-level CY25 MTD breakdown
    div_cy25 = {}
    for div, grp in cy25_mtd.groupby("Division"):
        div_cy25[div] = {
            "inflow":  int(grp["Repair Order#"].nunique()),
            "labour":  float(grp["Net Taxable Labor Amount"].sum()),
            "spares":  float(grp["Net Taxable Parts Amt"].sum()),
        }

    return {
        "rows": rows_out.to_dict(orient="records"),
        "summary": {
            "inflow":  int(gdf["Repair Order#"].nunique()) if "Repair Order#" in gdf.columns else 0,
            "labour":  indian_format(labour),
            "spares":  indian_format(spares),
            "total":   indian_format(labour + spares),
        },
        "mtd": {
            "label":         f"{curr_month_name} 2026 (1-{mtd_day})",
            "cy26_inflow":   int(gdf["Repair Order#"].nunique()) if "Repair Order#" in gdf.columns else 0,
            "cy25_inflow":   cy25_mtd_inflow,
            "inflow_pct":    pct(int(gdf["Repair Order#"].nunique()) if "Repair Order#" in gdf.columns else 0, cy25_mtd_inflow),
            "cy26_labour":   indian_format(labour),
            "cy25_labour":   indian_format(cy25_mtd_labour),
            "labour_pct":    pct(labour, cy25_mtd_labour),
            "cy26_spares":   indian_format(spares),
            "cy25_spares":   indian_format(cy25_mtd_spares),
            "spares_pct":    pct(spares, cy25_mtd_spares),
            "cy26_total":    indian_format(labour + spares),
            "cy25_total":    indian_format(cy25_mtd_labour + cy25_mtd_spares),
            "total_pct":     pct(labour + spares, cy25_mtd_labour + cy25_mtd_spares),
            "mtd_day":       mtd_day,
        },
        "div_cy25": div_cy25,
        "div_cy26": div_cy26,
        "mtd_label": f"Mar 2026 (1-{mtd_day}) vs CY25 Same Period",
        "month": f"March {curr_year}",
    }


@app.get("/", response_class=HTMLResponse)
def dashboard():
    return HTMLResponse("""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>Renault Service Revenue Dashboard v6</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }

body {
    font-family: Arial, sans-serif;
    background: linear-gradient(135deg, #4f6bdc, #7b4fdc);
    min-height: 100vh;
    padding: 24px;
}

.container {
    background: #fff;
    border-radius: 16px;
    padding: 28px 32px;
    max-width: 1400px;
    margin: 0 auto;
}

/* ---------- HEADER ---------- */
.header { display: flex; align-items: center; justify-content: space-between; margin-bottom: 20px; }
.header h2 { font-size: 22px; color: #222; }

/* ---------- TABS ---------- */
.tabs { display: flex; gap: 4px; margin-bottom: 24px; border-bottom: 2px solid #e5e7eb; }
.tab-btn {
    padding: 10px 24px;
    border: none;
    background: none;
    font-size: 14px;
    font-weight: 600;
    color: #888;
    cursor: pointer;
    border-bottom: 3px solid transparent;
    margin-bottom: -2px;
    border-radius: 6px 6px 0 0;
    transition: all .2s;
}
.tab-btn:hover { color: #4f6bdc; background: #f4f6ff; }
.tab-btn.active { color: #4f6bdc; border-bottom-color: #4f6bdc; background: #f4f6ff; }

.tab-page { display: none; }
.tab-page.active { display: block; }

/* ---------- CARDS ---------- */
.cards { display: flex; gap: 16px; margin-bottom: 24px; }
.card {
    flex: 1; background: #f4f6ff; border-radius: 10px;
    padding: 18px 16px; text-align: center;
}
.card .label {
    font-size: 12px; color: #777; font-weight: 600;
    text-transform: uppercase; letter-spacing: .5px; margin-bottom: 8px;
}
.card .value { font-size: 22px; font-weight: 700; color: #4f6bdc; }

/* ---------- FILTER BAR ---------- */
.filter-bar {
    display: flex; flex-wrap: wrap; gap: 12px; align-items: flex-end;
    background: #f7f8ff; border-radius: 10px; padding: 16px; margin-bottom: 20px;
}
.fg { display: flex; flex-direction: column; gap: 4px; flex: 1; min-width: 150px; position: relative; }
.fg label {
    font-size: 11px; font-weight: 700; color: #555;
    text-transform: uppercase; letter-spacing: .5px;
}
.fg select#cy {
    width: 100%; padding: 7px 10px; border: 1px solid #ccc; border-radius: 6px;
    font-size: 13px; background: #fff; cursor: pointer; color: #333;
    appearance: auto; height: 34px;
}
.fg select#cy:focus { outline: 2px solid #4f6bdc; border-color: transparent; }

/* ---------- CUSTOM MULTI-SELECT ---------- */
.custom-select { position: relative; width: 100%; }
.cs-face {
    width: 100%; height: 34px; padding: 0 28px 0 10px; border: 1px solid #ccc;
    border-radius: 6px; font-size: 13px;
    background: #fff url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6'%3E%3Cpath d='M0 0l5 6 5-6z' fill='%23666'/%3E%3C/svg%3E") no-repeat right 10px center;
    cursor: pointer; text-align: left; color: #333; white-space: nowrap;
    overflow: hidden; text-overflow: ellipsis; display: flex; align-items: center; user-select: none;
}
.cs-face:hover { border-color: #aaa; }
.cs-face.open { outline: 2px solid #4f6bdc; border-color: transparent; }
.cs-panel {
    display: none; position: absolute; top: calc(100% + 4px); left: 0;
    width: 100%; min-width: 200px; background: #fff; border: 1px solid #ccc;
    border-radius: 6px; box-shadow: 0 4px 16px rgba(0,0,0,.15);
    z-index: 999; max-height: 240px; overflow: hidden; flex-direction: column;
}
.cs-panel.open { display: flex; }
.cs-search { padding: 8px 8px 4px; border-bottom: 1px solid #eee; flex-shrink: 0; }
.cs-search input {
    width: 100%; padding: 5px 8px; border: 1px solid #ccc;
    border-radius: 4px; font-size: 12px; outline: none;
}
.cs-search input:focus { border-color: #4f6bdc; }
.cs-list { overflow-y: auto; flex: 1; padding: 4px 0; }
.cs-item { display: flex; align-items: center; gap: 8px; padding: 6px 10px; font-size: 13px; cursor: pointer; color: #333; }
.cs-item:hover { background: #f0f3ff; }
.cs-item input[type=checkbox] { accent-color: #4f6bdc; cursor: pointer; }
.cs-item.checked { background: #eef1ff; font-weight: 600; }
.cs-footer { border-top: 1px solid #eee; padding: 5px 10px; display: flex; justify-content: flex-end; flex-shrink: 0; }
.cs-clear { font-size: 11px; color: #4f6bdc; cursor: pointer; text-decoration: underline; }
/* Select All / Deselect All bar */
.cs-sel-all {
    display: flex;
    gap: 6px;
    padding: 6px 8px;
    border-bottom: 1px solid #eee;
    flex-shrink: 0;
}
.cs-sa-btn {
    flex: 1;
    padding: 4px 0;
    font-size: 11px;
    font-weight: 600;
    border: 1px solid #ccc;
    border-radius: 4px;
    background: #f4f6ff;
    color: #4f6bdc;
    cursor: pointer;
    white-space: nowrap;
}
.cs-sa-btn:hover { background: #4f6bdc; color: #fff; border-color: #4f6bdc; }


/* ---------- BUTTONS ---------- */
.apply-btn {
    padding: 0 26px; height: 34px; background: #4f6bdc; color: #fff;
    border: none; border-radius: 6px; font-size: 14px; font-weight: 700;
    cursor: pointer; align-self: flex-end; white-space: nowrap;
}
.apply-btn:hover { background: #3a56c5; }
.export-btn {
    padding: 0 20px; height: 34px; background: #17a34a; color: #fff;
    border: none; border-radius: 6px; font-size: 14px; font-weight: 700;
    cursor: pointer; align-self: flex-end; white-space: nowrap;
}
.export-btn:hover { background: #15803d; }
.export-btn:disabled { background: #86c9a3; cursor: not-allowed; }
.reset-btn {
    padding: 0 18px;
    height: 34px;
    background: #fff;
    color: #ef4444;
    border: 1.5px solid #ef4444;
    border-radius: 6px;
    font-size: 14px;
    font-weight: 700;
    cursor: pointer;
    align-self: flex-end;
    white-space: nowrap;
}
.reset-btn:hover { background: #ef4444; color: #fff; }


/* ---------- TABLE (Page 1) ---------- */
.table-wrap { overflow-x: auto; }
table { width: 100%; border-collapse: collapse; font-size: 13px; }
thead tr { background: #4f6bdc; color: #fff; }
th { padding: 10px 12px; text-align: left; white-space: nowrap; }
td { padding: 8px 12px; border-bottom: 1px solid #eee; white-space: nowrap; }
tbody tr:hover { background: #f4f6ff; }
.empty-msg { text-align: center; color: #999; padding: 24px; }

/* ---------- COMPARISON TABLE (Page 2) ---------- */
.comp-section { margin-bottom: 36px; }
.comp-section h3 {
    font-size: 15px; font-weight: 700; color: #4f6bdc;
    margin-bottom: 12px; padding-bottom: 6px;
    border-bottom: 2px solid #e5e7eb;
}
.comp-table { width: 100%; border-collapse: collapse; font-size: 13px; }
.comp-table thead tr { background: #4f6bdc; color: #fff; }
.comp-table th { padding: 10px 14px; text-align: center; white-space: nowrap; }
.comp-table th:first-child { text-align: left; }
.comp-table td { padding: 9px 14px; border-bottom: 1px solid #eee; text-align: right; white-space: nowrap; }
.comp-table td:first-child { text-align: left; font-weight: 600; color: #333; }
.comp-table tbody tr:hover { background: #f4f6ff; }
.comp-table tfoot tr { background: #eef1ff; font-weight: 700; }
.comp-table tfoot td { padding: 10px 14px; border-top: 2px solid #4f6bdc; }

/* PCT badges */
.pct,.pct-up,.pct-down,.pct-flat { display: inline-block; padding: 2px 8px; border-radius: 20px; font-size: 12px; font-weight: 700; }
.pct.up,  .pct-up   { background: #dcfce7; color: #15803d; }
.pct.down,.pct-down { background: #fee2e2; color: #dc2626; }
.pct.flat,.pct-flat { background: #f3f4f6; color: #6b7280; }

.comp-loading { text-align: center; color: #999; padding: 40px; font-size: 14px; }

/* ---------- DIVISION-MONTH PIVOT TABLE (Page 3) ---------- */
.pivot-wrap { overflow-x: auto; margin-bottom: 36px; }
.pivot-section h3 {
    font-size: 15px; font-weight: 700; color: #4f6bdc;
    margin-bottom: 12px; padding-bottom: 6px;
    border-bottom: 2px solid #e5e7eb;
}
.pivot-table { border-collapse: collapse; font-size: 11px; width: 100%; }
.pivot-table thead tr:first-child th {
    background: #4f6bdc; color: #fff; padding: 7px 8px;
    white-space: nowrap; text-align: center;
    border: 1px solid #3a56c5;
}
.pivot-table thead tr:first-child th:first-child { text-align: left; min-width: 80px; }
.pivot-table thead tr:nth-child(2) th {
    padding: 4px 6px; font-size: 10px; text-align: center;
    border: 1px solid #3a56c5; white-space: nowrap;
}
/* CY24 sub-header */
.pivot-table thead tr:nth-child(2) th:nth-child(3n+1) { background: #5a7ae8; color:#fff; }
/* CY25 sub-header */
.pivot-table thead tr:nth-child(2) th:nth-child(3n+2) { background: #4f6bdc; color:#fff; }
/* Growth% sub-header */
.pivot-table thead tr:nth-child(2) th:nth-child(3n)   { background: #2e4baa; color:#fff; }

.pivot-table td {
    padding: 6px 8px; border: 1px solid #e5e7eb;
    text-align: right; white-space: nowrap; font-size: 11px;
}
.pivot-table td:first-child {
    text-align: left; font-weight: 700; color: #4f6bdc;
    background: #f0f3ff; white-space: nowrap;
    position: sticky; left: 0; z-index: 1; min-width: 80px;
}
.pivot-table tbody tr:hover td            { background: #f4f6ff; }
.pivot-table tbody tr:hover td:first-child { background: #e8ecff; }
.pivot-table tfoot td {
    background: #dde3f8; font-weight: 700; color: #2a3e8c;
    border: 1px solid #b0bdee; border-top: 2px solid #4f6bdc;
    padding: 7px 8px; text-align: right;
}
.pivot-table tfoot td:first-child { text-align: left; position: sticky; left: 0; }
/* ---------- ONE PAGER TABLE (Page 4) ---------- */
.op-wrap { overflow-x: auto; margin-top: 8px; }
.op-table { width: 100%; border-collapse: collapse; font-size: 12px; }
.op-table th {
    background: #4f6bdc; color: #fff;
    padding: 9px 12px; text-align: center;
    border: 1px solid #3a56c5; white-space: nowrap;
}
.op-table th.desc-hdr { text-align: left; min-width: 220px; background: #2e4baa; }
.op-table th.div-hdr  { min-width: 80px; }
.op-table th.total-hdr{ background: #1e3a8a; }
.op-table th.avg-hdr  { background: #1e3a8a; }

.op-table td {
    padding: 8px 12px; border: 1px solid #e5e7eb;
    text-align: right; white-space: nowrap; font-size: 12px;
}
.op-table td.desc-cell {
    text-align: left; font-weight: 600; color: #222;
    background: #f8f9ff; position: sticky; left: 0; z-index: 1;
}
.op-table td.cat-sep {
    text-align: left; font-weight: 700; color: #fff;
    background: #4f6bdc; padding: 6px 12px; font-size: 11px;
    letter-spacing: .5px; text-transform: uppercase;
}
.op-table td.total-cell { font-weight: 700; background: #eef1ff; color: #2a3e8c; }
.op-table td.avg-cell   { font-weight: 700; background: #f0fdf4; color: #15803d; }
.op-table tbody tr:hover td { background: #f4f6ff; }
.op-table tbody tr:hover td.desc-cell { background: #e8ecff; }
.op-table tbody tr:hover td.total-cell { background: #dde3f8; }
.op-table tbody tr:hover td.avg-cell   { background: #dcfce7; }
.op-table .section-hdr td {
    background: #e8ecff; font-weight: 700; color: #4f6bdc;
    border-top: 2px solid #4f6bdc; font-size: 11px;
    text-transform: uppercase; letter-spacing: .4px;
}


@keyframes pulse {
  0%, 100% { opacity: 1; transform: scale(1); }
  50%       { opacity: .4; transform: scale(1.3); }
}
</style>
</head>

<body>
<div class="container">

  <!-- HEADER -->
  <div class="header">
    <h2>Renault Service Revenue Dashboard</h2>
  </div>

  <!-- TABS -->
  <div class="tabs">
    <button class="tab-btn active" onclick="switchTab('page1', this)">Invoice Data</button>
    <button class="tab-btn"        onclick="switchTab('page2', this)">CY Comparison</button>
    <button class="tab-btn"        onclick="switchTab('page3', this)">Division-Month</button>
    <button class="tab-btn"        onclick="switchTab('page4', this)">One Pager Report</button>
    <button class="tab-btn"        onclick="switchTab('page5', this)">Current Month (Mar)</button>
    <button class="tab-btn"        onclick="switchTab('page6', this)">Division-Month CY26</button>
  </div>

  <!-- ==================== PAGE 1 ==================== -->
  <div id="page1" class="tab-page active">

    <!-- CARDS -->
    <div class="cards">
      <div class="card"><div class="label">Inflow CY24</div><div class="value" id="inflow24">…</div></div>
      <div class="card"><div class="label">Inflow CY25</div><div class="value" id="inflow25">…</div></div>
      <div class="card"><div class="label">Inflow CY26</div><div class="value" id="inflow26">…</div></div>
      <div class="card"><div class="label">Total Labour</div><div class="value" id="labour">…</div></div>
      <div class="card"><div class="label">Total Spares</div><div class="value" id="spares">…</div></div>
      <div class="card" style="border: 2px solid #4f6bdc;"><div class="label">Total Revenue</div><div class="value" id="total-rev">…</div></div>
    </div>

    <!-- FILTER BAR -->
    <div class="filter-bar">
      <div class="fg">
        <label>Calendar Year</label>
        <select id="cy">
          <option value="">All CY</option>
          <option value="CY24">CY24</option>
          <option value="CY25">CY25</option>
          <option value="CY26">CY26</option>
        </select>
      </div>
      <div class="fg"><label>Division</label>
        <div class="custom-select" id="cs-division" data-key="division" data-placeholder="All Divisions"></div>
      </div>
      <div class="fg"><label>Service Type</label>
        <div class="custom-select" id="cs-service" data-key="service" data-placeholder="All Service Types"></div>
      </div>
      <div class="fg"><label>SA Name</label>
        <div class="custom-select" id="cs-sa" data-key="sa" data-placeholder="All SAs"></div>
      </div>
      <div class="fg"><label>Invoice Type</label>
        <div class="custom-select" id="cs-invoice" data-key="invoice" data-placeholder="All Invoice Types"></div>
      </div>
      <div class="fg"><label>Model</label>
        <div class="custom-select" id="cs-model" data-key="model" data-placeholder="All Models"></div>
      </div>
      <div class="fg"><label>Month</label>
        <div class="custom-select" id="cs-month" data-key="month" data-placeholder="All Months"></div>
      </div>
      <button class="apply-btn"  onclick="applyFilters()">Apply</button>
      <button class="export-btn" onclick="exportExcel()">Export Excel</button>
      <button class="reset-btn"  onclick="resetPage1()">Reset</button>
    </div>

    <!-- TABLE -->
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Division</th><th>RO #</th><th>SA Name</th>
            <th>Vehicle</th><th>Service Type</th>
            <th>Labour (₹)</th><th>Spares (₹)</th>
            <th style="background:#3a56c5;">Total Revenue (₹)</th>
          </tr>
        </thead>
        <tbody id="tbody">
          <tr><td colspan="8" class="empty-msg">Loading…</td></tr>
        </tbody>
      </table>
    </div>

  </div><!-- /page1 -->

  <!-- ==================== PAGE 2 ==================== -->
  <div id="page2" class="tab-page">

    <!-- FILTER BAR (Comparison) -->
    <div class="filter-bar">
      <div class="fg">
        <label>Division</label>
        <div class="custom-select" id="cs2-division" data-key="c-division" data-placeholder="All Divisions"></div>
      </div>
      <div class="fg">
        <label>Service Type</label>
        <div class="custom-select" id="cs2-service" data-key="c-service" data-placeholder="All Service Types"></div>
      </div>
      <div class="fg">
        <label>SA Name</label>
        <div class="custom-select" id="cs2-sa" data-key="c-sa" data-placeholder="All SAs"></div>
      </div>
      <div class="fg">
        <label>Invoice Type</label>
        <div class="custom-select" id="cs2-invoice" data-key="c-invoice" data-placeholder="All Invoice Types"></div>
      </div>
      <div class="fg">
        <label>Model</label>
        <div class="custom-select" id="cs2-model" data-key="c-model" data-placeholder="All Models"></div>
      </div>
      <button class="apply-btn"  onclick="applyComparison()">Apply</button>
      <button class="export-btn" id="export-btn-p2" onclick="exportPage2()">Export Excel</button>
      <button class="reset-btn"  onclick="resetPage2()">Reset</button>
    </div>

    <div id="comp-content">
      <div class="comp-loading">Click the tab to load comparison data…</div>
    </div>

  </div><!-- /page2 -->

  <!-- ==================== PAGE 3 ==================== -->
  <div id="page3" class="tab-page">

    <!-- FILTER BAR (Division-Month) -->
    <div class="filter-bar">
      <div class="fg">
        <label>Division</label>
        <div class="custom-select" id="cs3-division" data-key="d-division" data-placeholder="All Divisions"></div>
      </div>
      <div class="fg">
        <label>Service Type</label>
        <div class="custom-select" id="cs3-service" data-key="d-service" data-placeholder="All Service Types"></div>
      </div>
      <div class="fg">
        <label>SA Name</label>
        <div class="custom-select" id="cs3-sa" data-key="d-sa" data-placeholder="All SAs"></div>
      </div>
      <div class="fg">
        <label>Invoice Type</label>
        <div class="custom-select" id="cs3-invoice" data-key="d-invoice" data-placeholder="All Invoice Types"></div>
      </div>
      <div class="fg">
        <label>Model</label>
        <div class="custom-select" id="cs3-model" data-key="d-model" data-placeholder="All Models"></div>
      </div>
      <button class="apply-btn"  onclick="applyDivMonth()">Apply</button>
      <button class="export-btn" id="export-btn-p3" onclick="exportPage3()">Export Excel</button>
      <button class="reset-btn"  onclick="resetPage3()">Reset</button>
    </div>

    <div id="divmonth-content">
      <div class="comp-loading">Click the tab to load data…</div>
    </div>

  </div><!-- /page3 -->

  <!-- ==================== PAGE 4 ==================== -->
  <div id="page4" class="tab-page">

    <!-- FILTER BAR -->
    <div class="filter-bar">
      <div class="fg">
        <label>Calendar Year</label>
        <select id="op-cy">
          <option value="">All CY</option>
          <option value="CY24">CY24</option>
          <option value="CY25">CY25</option>
          <option value="CY26">CY26</option>
        </select>
      </div>
      <div class="fg"><label>Division</label>
        <div class="custom-select" id="cs4-division" data-key="op-division" data-placeholder="All Divisions"></div>
      </div>
      <div class="fg"><label>Service Type</label>
        <div class="custom-select" id="cs4-service" data-key="op-service" data-placeholder="All Service Types"></div>
      </div>
      <div class="fg"><label>SA Name</label>
        <div class="custom-select" id="cs4-sa" data-key="op-sa" data-placeholder="All SAs"></div>
      </div>
      <div class="fg"><label>Invoice Type</label>
        <div class="custom-select" id="cs4-invoice" data-key="op-invoice" data-placeholder="All Invoice Types"></div>
      </div>
      <div class="fg"><label>Model</label>
        <div class="custom-select" id="cs4-model" data-key="op-model" data-placeholder="All Models"></div>
      </div>
      <div class="fg"><label>Month</label>
        <div class="custom-select" id="cs4-month" data-key="op-month" data-placeholder="All Months"></div>
      </div>
      <button class="apply-btn"  onclick="applyOnePager()">Apply</button>
      <button class="export-btn" id="export-btn-p4" onclick="exportPage4()">Export Excel</button>
      <button class="reset-btn"  onclick="resetPage4()">Reset</button>
    </div>

    <div id="op-content">
      <div class="comp-loading">Click the tab to load One Pager Report…</div>
    </div>

  </div><!-- /page4 -->

  <!-- ==================== PAGE 5 — CURRENT MONTH ==================== -->
  <div id="page5" class="tab-page">

    <!-- FILTER BAR -->
    <div class="filter-bar">
      <div class="fg"><label>Division</label>
        <div class="custom-select" id="cs5-division" data-key="cm-division" data-placeholder="All Divisions"></div>
      </div>
      <div class="fg"><label>Service Type</label>
        <div class="custom-select" id="cs5-service" data-key="cm-service" data-placeholder="All Service Types"></div>
      </div>
      <div class="fg"><label>SA Name</label>
        <div class="custom-select" id="cs5-sa" data-key="cm-sa" data-placeholder="All SAs"></div>
      </div>
      <div class="fg"><label>Model</label>
        <div class="custom-select" id="cs5-model" data-key="cm-model" data-placeholder="All Models"></div>
      </div>
      <button class="apply-btn"  onclick="applyCurrentMonth()">Apply</button>
      <button class="export-btn" id="export-btn-p5" onclick="exportPage5()">Export Excel</button>
      <button class="reset-btn"  onclick="resetPage5()">Reset</button>
    </div>

    <!-- LIVE CARDS -->
    <div class="cards" style="margin-bottom:20px;">
      <div class="card" style="border:2px solid #17a34a;">
        <div class="label">Current Month</div>
        <div class="value" style="color:#17a34a;" id="cm-month-label">March 2026</div>
      </div>
      <div class="card"><div class="label">Total Inflow</div><div class="value" id="cm-inflow">…</div></div>
      <div class="card"><div class="label">Labour Revenue</div><div class="value" id="cm-labour">…</div></div>
      <div class="card"><div class="label">Spares Revenue</div><div class="value" id="cm-spares">…</div></div>
      <div class="card" style="border:2px solid #4f6bdc;"><div class="label">Total Revenue</div><div class="value" id="cm-total">…</div></div>
    </div>

    <!-- LIVE BADGE -->
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:14px;">
      <span style="display:inline-flex;align-items:center;gap:6px;background:#dcfce7;color:#15803d;padding:5px 14px;border-radius:20px;font-size:12px;font-weight:700;">
        <span style="width:8px;height:8px;background:#15803d;border-radius:50%;display:inline-block;animation:pulse 1.5s infinite;"></span>
        LIVE — Google Sheets
      </span>
      <button onclick="loadCurrentMonth()" style="background:none;border:1px solid #ccc;border-radius:6px;padding:4px 12px;font-size:12px;cursor:pointer;color:#555;">Refresh</button>
      <span id="cm-last-updated" style="font-size:11px;color:#999;"></span>
    </div>

    <!-- MTD COMPARISON TABLE -->
    <div id="cm-mtd-wrap" style="margin-bottom:20px;overflow-x:auto;display:none;">
      <h3 style="font-size:14px;font-weight:700;color:#4f6bdc;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid #e5e7eb;">
        MTD Comparison — <span id="cm-mtd-label">Current Month</span> vs Same Period Last Year (CY25)
      </h3>
      <table class="comp-table" style="min-width:500px;">
        <thead>
          <tr>
            <th style="text-align:left;min-width:160px;">Metric</th>
            <th id="cm-mtd-cy25-hdr">CY25 MTD</th>
            <th id="cm-mtd-cy26-hdr">CY26 MTD</th>
            <th>Growth %</th>
          </tr>
        </thead>
        <tbody id="cm-mtd-tbody">
          <tr><td colspan="4" class="empty-msg">Loading MTD data&#8230;</td></tr>
        </tbody>
      </table>
    </div>

    <!-- DIVISION SUMMARY TABLE -->
    <div id="cm-div-wrap" style="margin-bottom:20px;overflow-x:auto;display:none;">
      <h3 style="font-size:14px;font-weight:700;color:#4f6bdc;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid #e5e7eb;">
        Division-wise Summary — <span id="cm-div-month-label">Current Month</span>
      </h3>
      <table class="comp-table" style="min-width:900px;">
        <thead>
          <tr>
            <th rowspan=2 style="text-align:left;min-width:80px;">Division</th>
            <th colspan=3 style="background:#4f6bdc;">Inflow (ROs)</th>
            <th colspan=3 style="background:#4f6bdc;">Labour (&#8377;)</th>
            <th colspan=3 style="background:#4f6bdc;">Spares (&#8377;)</th>
            <th colspan=3 style="background:#1e3a8a;">Total Revenue (&#8377;)</th>
          </tr>
          <tr>
            <th style="background:#5a7ae8;font-size:11px;">CY25 MTD</th>
            <th style="background:#17a34a;font-size:11px;">CY26 MTD</th>
            <th style="background:#2e4baa;font-size:11px;">Growth%</th>
            <th style="background:#5a7ae8;font-size:11px;">CY25 MTD</th>
            <th style="background:#17a34a;font-size:11px;">CY26 MTD</th>
            <th style="background:#2e4baa;font-size:11px;">Growth%</th>
            <th style="background:#5a7ae8;font-size:11px;">CY25 MTD</th>
            <th style="background:#17a34a;font-size:11px;">CY26 MTD</th>
            <th style="background:#2e4baa;font-size:11px;">Growth%</th>
            <th style="background:#5a7ae8;font-size:11px;">CY25 MTD</th>
            <th style="background:#17a34a;font-size:11px;">CY26 MTD</th>
            <th style="background:#2e4baa;font-size:11px;">Growth%</th>
          </tr>
        </thead>
        <tbody id="cm-div-tbody">
          <tr><td colspan="13" class="empty-msg">Loading&#8230;</td></tr>
        </tbody>
      </table>
    </div>

    <!-- DETAIL TABLE -->
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Division</th><th>RO #</th><th>SA Name</th>
            <th>Vehicle</th><th>Service Type</th>
            <th>Labour (&#8377;)</th><th>Spares (&#8377;)</th>
            <th style="background:#3a56c5;">Total Revenue (&#8377;)</th>
          </tr>
        </thead>
        <tbody id="cm-tbody">
          <tr><td colspan="8" class="empty-msg">Click the tab to load live data…</td></tr>
        </tbody>
      </table>
    </div>

  </div><!-- /page5 -->

  <!-- ==================== PAGE 6 — DIVISION-MONTH CY25 vs CY26 ==================== -->
  <div id="page6" class="tab-page">

    <div class="filter-bar">
      <div class="fg"><label>Division</label>
        <div class="custom-select" id="cs6-division" data-key="d6-division" data-placeholder="All Divisions"></div>
      </div>
      <div class="fg"><label>Service Type</label>
        <div class="custom-select" id="cs6-service" data-key="d6-service" data-placeholder="All Service Types"></div>
      </div>
      <div class="fg"><label>SA Name</label>
        <div class="custom-select" id="cs6-sa" data-key="d6-sa" data-placeholder="All SAs"></div>
      </div>
      <div class="fg"><label>Invoice Type</label>
        <div class="custom-select" id="cs6-invoice" data-key="d6-invoice" data-placeholder="All Invoice Types"></div>
      </div>
      <div class="fg"><label>Model</label>
        <div class="custom-select" id="cs6-model" data-key="d6-model" data-placeholder="All Models"></div>
      </div>
      <button class="apply-btn"  onclick="applyDivMonth26()">Apply</button>
      <button class="export-btn" id="export-btn-p6" onclick="exportPage6()">Export Excel</button>
      <button class="reset-btn"  onclick="resetPage6()">Reset</button>
    </div>

    <div id="divmonth26-content">
      <div class="comp-loading">Click the tab to load CY25 vs CY26 data&#8230;</div>
    </div>

  </div><!-- /page6 -->

</div><!-- /container -->

<script>
/* ====================================================
   TAB SWITCHING
   ==================================================== */
let compLoaded    = false;
let divMonLoaded  = false;
let dm26Loaded    = false;
let opLoaded      = false;

function switchTab(pageId, btn) {
    document.querySelectorAll(".tab-page").forEach(p => p.classList.remove("active"));
    document.querySelectorAll(".tab-btn").forEach(b => b.classList.remove("active"));
    document.getElementById(pageId).classList.add("active");
    btn.classList.add("active");
    if (pageId === "page2" && !compLoaded)   { loadComparison(); compLoaded = true; }
    if (pageId === "page3" && !divMonLoaded) { loadDivMonth(); divMonLoaded = true; }
    if (pageId === "page4") { loadOnePager(); }
    if (pageId === "page5") { loadCurrentMonth(); }
    if (pageId === "page6" && !dm26Loaded) { loadDivMonth26(); dm26Loaded = true; }
}

/* ====================================================
   CUSTOM MULTI-SELECT COMPONENT
   ==================================================== */
const selections = {};

function buildCustomSelect(container) {
    const key         = container.dataset.key;
    const placeholder = container.dataset.placeholder || "Select…";
    selections[key]   = new Set();

    container.innerHTML = `
      <div class="cs-face" id="face-${key}">${placeholder}</div>
      <div class="cs-panel" id="panel-${key}">
        <div class="cs-search"><input type="text" placeholder="Search…" id="search-${key}"></div>
        <div class="cs-sel-all">
          <button class="cs-sa-btn" onclick="selectAllVisible('${key}')">✔ Select All</button>
          <button class="cs-sa-btn" onclick="clearSelect('${key}')">✖ Deselect All</button>
        </div>
        <div class="cs-list"  id="list-${key}"></div>
      </div>`;

    document.getElementById("face-" + key).addEventListener("click", (e) => {
        e.stopPropagation(); togglePanel(key);
    });
    document.getElementById("search-" + key).addEventListener("input", (e) => {
        filterList(key, e.target.value.toLowerCase());
    });
}

function togglePanel(key) {
    document.querySelectorAll(".cs-panel.open").forEach(p => {
        if (p.id !== "panel-" + key) {
            p.classList.remove("open");
            document.getElementById("face-" + p.id.replace("panel-","")).classList.remove("open");
        }
    });
    const panel = document.getElementById("panel-" + key);
    const face  = document.getElementById("face-"  + key);
    panel.classList.toggle("open");
    face.classList.toggle("open", panel.classList.contains("open"));
}


function updateFace(key) {
    const face        = document.getElementById("face-" + key);
    const placeholder = document.querySelector(`[data-key="${key}"]`).dataset.placeholder;
    const sel         = selections[key];
    face.textContent  = sel.size === 0 ? placeholder : sel.size === 1 ? [...sel][0] : sel.size + " selected";
}

document.addEventListener("click", () => {
    document.querySelectorAll(".cs-panel.open").forEach(p => {
        p.classList.remove("open");
        document.getElementById("face-" + p.id.replace("panel-","")).classList.remove("open");
    });
});

/* ====================================================
   API HELPERS
   ==================================================== */
function buildParams() {
    const p = new URLSearchParams();
    const cy = document.getElementById("cy").value;
    if (cy) p.append("cy", cy);
    ["division","service","sa","invoice","model","month"].forEach(key => {
        selections[key].forEach(v => p.append(key, v));
    });
    return p.toString();
}

function fmt(n) {
    return Number(Math.round(n || 0)).toLocaleString("en-IN");
}

function pctBadge(val) {
    if (val === null || val === undefined) return '<span class=pct-flat>N/A</span>';
    const cls  = val > 0 ? "pct-up" : val < 0 ? "pct-down" : "pct-flat";
    const sign = val > 0 ? "\u25b2 +" : val < 0 ? "\u25bc " : "";
    return "<span class=" + cls + ">" + sign + val.toFixed(1) + "%</span>";
}

/* ====================================================
   PAGE 1 — FILTERS, CARDS, TABLE
   ==================================================== */
async function loadFilters() {
    try {
        const data = await fetch("/filters").then(r => r.json());
        // Page 1
        fillCustomSelect("division", data.division);
        fillCustomSelect("service",  data.service);
        fillCustomSelect("sa",       data.sa);
        fillCustomSelect("invoice",  data.invoice);
        fillCustomSelect("month",    data.month);
        fillCustomSelect("model",    data.model);
        // Page 2 — CY Comparison
        fillCustomSelect("c-division", data.division);
        fillCustomSelect("c-service",  data.service);
        fillCustomSelect("c-sa",       data.sa);
        fillCustomSelect("c-invoice",  data.invoice);
        fillCustomSelect("c-model",    data.model);
        // Page 3 — Division-Month
        fillCustomSelect("d-division", data.division);
        fillCustomSelect("d-service",  data.service);
        fillCustomSelect("d-sa",       data.sa);
        fillCustomSelect("d-invoice",  data.invoice);
        fillCustomSelect("d-model",    data.model);
        // Page 4 — One Pager
        fillCustomSelect("op-division", data.division);
        fillCustomSelect("op-service",  data.service);
        fillCustomSelect("op-sa",       data.sa);
        fillCustomSelect("op-invoice",  data.invoice);
        fillCustomSelect("op-model",    data.model);
        fillCustomSelect("op-month",    data.month);
        // Page 5 — Current Month
        fillCustomSelect("cm-division", data.division);
        fillCustomSelect("cm-service",  data.service);
        fillCustomSelect("cm-sa",       data.sa);
        fillCustomSelect("cm-model",    data.model);
        // Page 6 — Division-Month CY26
        fillCustomSelect("d6-division", data.division);
        fillCustomSelect("d6-service",  data.service);
        fillCustomSelect("d6-sa",       data.sa);
        fillCustomSelect("d6-invoice",  data.invoice);
        fillCustomSelect("d6-model",    data.model);
    } catch(e) {
        console.error("loadFilters failed:", e);
    }
}

function buildCompParams() {
    const p = new URLSearchParams();
    ["c-division","c-service","c-sa","c-invoice","c-model"].forEach(key => {
        selections[key].forEach(v => p.append(key.replace("c-",""), v));
    });
    return p.toString();
}

function applyComparison() {
    compLoaded = true;
    loadComparison();
}

async function loadCards() {
    try {
        const data = await fetch("/cards?" + buildParams()).then(r => r.json());
        document.getElementById("inflow24").innerText  = Number(data.inflow24).toLocaleString("en-IN");
        document.getElementById("inflow25").innerText  = Number(data.inflow25).toLocaleString("en-IN");
        document.getElementById("inflow26").innerText  = Number(data.inflow26||0).toLocaleString("en-IN");
        document.getElementById("labour").innerText    = data.labour;
        document.getElementById("spares").innerText    = data.spares;
        document.getElementById("total-rev").innerText = data.total;
    } catch(e) { console.error("Card load failed", e); }
}

async function loadTable() {
    const tbody = document.getElementById("tbody");
    tbody.innerHTML = '<tr><td colspan="8" class="empty-msg">Loading…</td></tr>';
    try {
        const data = await fetch("/table?" + buildParams()).then(r => r.json());
        if (!data.length) {
            tbody.innerHTML = '<tr><td colspan="8" class="empty-msg">No records found</td></tr>';
            return;
        }
        tbody.innerHTML = data.map(r => `
          <tr>
            <td>${r.Division ?? ""}</td>
            <td>${r["Repair Order#"] ?? ""}</td>
            <td>${r["SA Name"] ?? ""}</td>
            <td>${r["Vehicle Reg#"] ?? ""}</td>
            <td>${r["Service Type"] ?? ""}</td>
            <td>${fmt(r["Net Taxable Labor Amount"])}</td>
            <td>${fmt(r["Net Taxable Parts Amt"])}</td>
            <td style="font-weight:600;color:#4f6bdc;">${fmt(r["Total Revenue"])}</td>
          </tr>`).join("");
    } catch(e) {
        tbody.innerHTML = '<tr><td colspan="7" class="empty-msg">Error loading data</td></tr>';
        console.error("Table load failed", e);
    }
}

function applyFilters() { loadCards(); loadTable(); }

async function exportExcel() {
    const btn = document.querySelector(".export-btn");
    btn.disabled = true; btn.textContent = "Exporting...";
    try {
        const res  = await fetch("/export?" + buildParams());
        if (!res.ok) throw new Error("Export failed");
        const blob = await res.blob();
        const a = document.createElement("a");
        a.href = URL.createObjectURL(blob);
        a.download = "Renault_Invoice_Export.xlsx";
        document.body.appendChild(a); a.click();
        document.body.removeChild(a); URL.revokeObjectURL(a.href);
    } catch(e) { alert("Export failed: " + e.message); }
    finally { btn.disabled = false; btn.textContent = "Export Excel"; }
}

async function _triggerExport(url, filename, btnId) {
    const btn = document.getElementById(btnId);
    if (btn) { btn.disabled = true; btn.textContent = "Exporting..."; }
    try {
        const res = await fetch(url);
        if (!res.ok) throw new Error("Export failed: " + res.status);
        const blob = await res.blob();
        const a = document.createElement("a");
        a.href = URL.createObjectURL(blob);
        a.download = filename;
        document.body.appendChild(a); a.click();
        document.body.removeChild(a); URL.revokeObjectURL(a.href);
    } catch(e) { alert("Export failed: " + e.message); }
    finally { if (btn) { btn.disabled = false; btn.textContent = "Export Excel"; } }
}

function exportPage2() {
    _triggerExport("/export-comparison?" + buildCompParams(), "Renault_CY_Comparison_Export.xlsx", "export-btn-p2");
}

function exportPage3() {
    _triggerExport("/export-division-month?" + buildDivMonParams(), "Renault_DivisionMonth_Export.xlsx", "export-btn-p3");
}

function exportPage4() {
    _triggerExport("/export-one-pager?" + buildOpParams(), "Renault_OnePager_Export.xlsx", "export-btn-p4");
}

function exportPage5() {
    _triggerExport("/export-current-month?" + buildCmParams(), "Renault_CurrentMonth_Export.xlsx", "export-btn-p5");
}

function exportPage6() {
    _triggerExport("/export-division-month-cy26?" + buildDm26Params(), "Renault_DivMonth_CY26_Export.xlsx", "export-btn-p6");
}

/* ====================================================
   PAGE 2 — CY COMPARISON
   ==================================================== */
async function loadComparison() {
    const wrap = document.getElementById("comp-content");
    wrap.innerHTML = '<div class="comp-loading">Loading comparison data…</div>';
    try {
        const rows = await fetch("/comparison?" + buildCompParams()).then(r => r.json());
        if (!rows.length) {
            wrap.innerHTML = '<div class="comp-loading">No data available</div>'; return;
        }

        // Totals
        const tot = { i24:0, i25:0, l24:0, l25:0, s24:0, s25:0, t24:0, t25:0 };
        rows.forEach(r => {
            tot.i24 += r.inflow24; tot.i25 += r.inflow25;
            tot.l24 += r.labour24; tot.l25 += r.labour25;
            tot.s24 += r.spares24; tot.s25 += r.spares25;
            tot.t24 += r.total24;  tot.t25 += r.total25;
        });
        const totIPct = tot.i24 ? ((tot.i25-tot.i24)/tot.i24*100).toFixed(1) : null;
        const totLPct = tot.l24 ? ((tot.l25-tot.l24)/tot.l24*100).toFixed(1) : null;
        const totSPct = tot.s24 ? ((tot.s25-tot.s24)/tot.s24*100).toFixed(1) : null;
        const totTPct = tot.t24 ? ((tot.t25-tot.t24)/tot.t24*100).toFixed(1) : null;

        function buildTable(title, cy24key, cy25key, pctKey, isCurrency) {
            const f = v => isCurrency ? "₹ " + fmt(v) : Number(v).toLocaleString("en-IN");
            const rows_html = rows.map(r => `
              <tr>
                <td>${r.month}</td>
                <td>${f(r[cy24key])}</td>
                <td>${f(r[cy25key])}</td>
                <td>${pctBadge(r[pctKey])}</td>
              </tr>`).join("");

            const t24 = isCurrency ? tot.l24 : tot.i24;
            const t25 = isCurrency ? tot.l25 : tot.i25;
            // pick right totals
            let tot24, tot25, totPct;
            if (cy24key === "inflow24")  { tot24=tot.i24; tot25=tot.i25; totPct=totIPct; }
            else if (cy24key === "labour24") { tot24=tot.l24; tot25=tot.l25; totPct=totLPct; }
            else if (cy24key === "spares24") { tot24=tot.s24; tot25=tot.s25; totPct=totSPct; }
            else { tot24=tot.t24; tot25=tot.t25; totPct=totTPct; }

            return `
              <div class="comp-section">
                <h3>${title}</h3>
                <div class="table-wrap">
                  <table class="comp-table">
                    <thead>
                      <tr>
                        <th>Month</th>
                        <th>CY 2024</th>
                        <th>CY 2025</th>
                        <th>Difference %</th>
                      </tr>
                    </thead>
                    <tbody>${rows_html}</tbody>
                    <tfoot>
                      <tr>
                        <td>Total</td>
                        <td>${f(tot24)}</td>
                        <td>${f(tot25)}</td>
                        <td>${pctBadge(parseFloat(totPct))}</td>
                      </tr>
                    </tfoot>
                  </table>
                </div>
              </div>`;
        }

        wrap.innerHTML =
            buildTable("📥 Inflow Comparison (Unique ROs)", "inflow24", "inflow25", "inflow_pct", false) +
            buildTable("🔧 Labour Revenue Comparison",      "labour24", "labour25", "labour_pct", true)  +
            buildTable("🔩 Spares Revenue Comparison",      "spares24", "spares25", "spares_pct", true)  +
            buildTable("💰 Total Revenue Comparison",       "total24",  "total25",  "total_pct",  true);

    } catch(e) {
        wrap.innerHTML = '<div class="comp-loading">Error loading comparison data</div>';
        console.error("Comparison load failed", e);
    }
}

/* ====================================================
   PAGE 3 — DIVISION-MONTH PIVOT
   ==================================================== */
function buildDivMonParams() {
    const p = new URLSearchParams();
    ["d-division","d-service","d-sa","d-invoice","d-model"].forEach(key => {
        selections[key].forEach(v => p.append(key.replace("d-",""), v));
    });
    return p.toString();
}

function applyDivMonth() { divMonLoaded = true; loadDivMonth(); }

async function loadDivMonth() {
    const wrap = document.getElementById("divmonth-content");
    wrap.innerHTML = '<div class="comp-loading">Loading…</div>';
    try {
        const { divisions, months, rows } = await fetch("/division-month?" + buildDivMonParams()).then(r => r.json());

        if (!divisions.length) {
            wrap.innerHTML = '<div class="comp-loading">No data available</div>'; return;
        }

        // Index rows by division
        const byDiv = {};
        rows.forEach(r => byDiv[r.division] = r);

        function gPct(v24, v25) {
            if (!v24) return '<span class=pct-flat>N/A</span>';
            const p = ((v25 - v24) / v24 * 100);
            const cls  = p > 0 ? "pct-up" : p < 0 ? "pct-down" : "pct-flat";
            const sign = p > 0 ? "\u25b2 +" : p < 0 ? "\u25bc " : "";
            return "<span class=" + cls + ">" + sign + p.toFixed(1) + "%</span>";
        }

        function buildPivot(title, prefix, isCurrency) {
            const f = v => isCurrency
                ? fmt(v)
                : Number(v||0).toLocaleString("en-IN");

            // Header row 1: Division | Jan(span3) | Feb(span3) | … | Total(span3)
            let hdr1 = `<th rowspan="2" style="min-width:80px;">Division</th>`;
            months.forEach(m => hdr1 += `<th colspan="3">${m}</th>`);
            hdr1 += `<th colspan="3">Total</th>`;

            // Header row 2: CY24 | CY25 | Growth% per month + total
            let hdr2 = "";
            [...months, "Total"].forEach(() =>
                hdr2 += `<th>CY24</th><th>CY25</th><th>Growth%</th>`);

            // Data rows
            let body = "";
            const gTot24 = new Array(months.length).fill(0);
            const gTot25 = new Array(months.length).fill(0);

            divisions.forEach(div => {
                const r = byDiv[div];
                let cells = "";
                let rSum24 = 0, rSum25 = 0;
                months.forEach((m, mi) => {
                    const v24 = r ? (r[`${prefix}24_${m}`] || 0) : 0;
                    const v25 = r ? (r[`${prefix}25_${m}`] || 0) : 0;
                    rSum24 += v24; rSum25 += v25;
                    gTot24[mi] += v24; gTot25[mi] += v25;
                    cells += `<td>${f(v24)}</td><td>${f(v25)}</td><td>${gPct(v24,v25)}</td>`;
                });
                body += `<tr>
                    <td>${div}</td>
                    ${cells}
                    <td style="font-weight:700;">${f(rSum24)}</td>
                    <td style="font-weight:700;color:#4f6bdc;">${f(rSum25)}</td>
                    <td>${gPct(rSum24, rSum25)}</td>
                  </tr>`;
            });

            // Footer: Grand Total
            let footCells = "";
            let gSum24 = 0, gSum25 = 0;
            gTot24.forEach((v, i) => {
                gSum24 += v; gSum25 += gTot25[i];
                footCells += `<td>${f(v)}</td><td>${f(gTot25[i])}</td><td>${gPct(v, gTot25[i])}</td>`;
            });

            return `
              <div class="pivot-section comp-section">
                <h3>${title}</h3>
                <div class="pivot-wrap">
                  <table class="pivot-table">
                    <thead>
                      <tr>${hdr1}</tr>
                      <tr>${hdr2}</tr>
                    </thead>
                    <tbody>${body}</tbody>
                    <tfoot>
                      <tr>
                        <td>Grand Total</td>
                        ${footCells}
                        <td>${f(gSum24)}</td>
                        <td>${f(gSum25)}</td>
                        <td>${gPct(gSum24, gSum25)}</td>
                      </tr>
                    </tfoot>
                  </table>
                </div>
              </div>`;
        }

        wrap.innerHTML =
            buildPivot("📥 Inflow — Division × Month (CY24 vs CY25)",         "i", false) +
            buildPivot("🔧 Labour — Division × Month (CY24 vs CY25)",         "l", true)  +
            buildPivot("🔩 Spares — Division × Month (CY24 vs CY25)",         "s", true)  +
            buildPivot("💰 Total Revenue — Division × Month (CY24 vs CY25)",  "t", true);

    } catch(e) {
        wrap.innerHTML = '<div class="comp-loading">Error loading data</div>';
        console.error("DivMonth load failed", e);
    }
}


/* ====================================================
   PAGE 4 — ONE PAGER REPORT
   ==================================================== */
function buildOpParams() {
    var p = new URLSearchParams();
    var cy = document.getElementById("op-cy").value;
    if (cy) p.append("cy", cy);
    ["op-division","op-service","op-sa","op-invoice","op-model","op-month"].forEach(function(key) {
        (selections[key] || new Set()).forEach(function(v) { p.append(key.replace("op-",""), v); });
    });
    return p.toString();
}

function applyOnePager() { loadOnePager(); }

async function loadOnePager() {
    var wrap = document.getElementById("op-content");
    wrap.innerHTML = '<div class="comp-loading">Loading One Pager Report...</div>';
    try {
        var resp = await fetch("/one-pager?" + buildOpParams());
        var json = await resp.json();
        var divisions = json.divisions;
        var rows = json.rows;

        if (!rows || !rows.length) {
            wrap.innerHTML = '<div class="comp-loading">No data available. Check service type mapping.</div>';
            return;
        }

        var byDiv = {};
        rows.forEach(function(r) { byDiv[r.division] = r; });
        var totalRow = byDiv["TOTAL"] || {};
        var divCols  = divisions;

        function fi(v) { return Number(v||0).toLocaleString("en-IN"); }
        function fc(v) { return fmt(v||0); }
        function fn(r, key) { return (r && r[key]) ? r[key] : 0; }
        function favg(key, isCurr) {
            var vals = divCols.map(function(d){ return fn(byDiv[d],key); }).filter(function(v){ return v > 0; });
            if (!vals.length) return 0;
            var a = vals.reduce(function(a,b){ return a+b; }, 0) / vals.length;
            return isCurr ? fc(a) : fi(Math.round(a));
        }

        function rowCells(key, isCurr) {
            var f = isCurr ? fc : fi;
            var cells = divCols.map(function(d){ return '<td>' + f(fn(byDiv[d],key)) + '</td>'; }).join('');
            cells += '<td class="total-cell">' + f(fn(totalRow,key)) + '</td>';
            cells += '<td class="avg-cell">'   + favg(key,isCurr)    + '</td>';
            return cells;
        }

        function secHdr(t, n) {
            return '<tr class="section-hdr"><td class="desc-cell" colspan="' + n + '">' + t + '</td></tr>';
        }
        function dataRow(label, key, isCurr, extraStyle) {
            var style = extraStyle ? ' style="' + extraStyle + '"' : '';
            return '<tr' + style + '><td class="desc-cell">' + label + '</td>' + rowCells(key,isCurr) + '</tr>';
        }

        var n = divCols.length + 3;
        var colHdrs = divCols.map(function(d){ return '<th class="div-hdr">' + d + '</th>'; }).join('') +
                      '<th class="total-hdr">TOTAL</th><th class="avg-hdr">AVERAGE</th>';

        var tbody =
            secHdr('&#128203; RO INFLOW', n) +
            dataRow('Total No of RO Billed (PDI + Acc)', 'total_billed_ro', false) +
            dataRow('&nbsp;&nbsp;&nbsp;PDI RO',           'pdi_ro',          false) +
            dataRow('&nbsp;&nbsp;&nbsp;Accessories',       'acc_ro',          false) +
            secHdr('&#128295; MECHANICAL', n) +
            dataRow('&nbsp;&nbsp;&nbsp;1 Free Service',   'fs1_ro', false) +
            dataRow('&nbsp;&nbsp;&nbsp;2 Free Service',   'fs2_ro', false) +
            dataRow('&nbsp;&nbsp;&nbsp;3 Free Service',   'fs3_ro', false) +
            dataRow('&nbsp;&nbsp;&nbsp;Paid Services RO', 'ps_ro',  false) +
            dataRow('&nbsp;&nbsp;&nbsp;Running Repair',   'rr_ro',  false) +
            dataRow('Mechanical RO', 'mech_ro', false, 'font-weight:700;') +
            secHdr('&#128297; BODY &amp; PAINT', n) +
            dataRow('Body &amp; Paint RO', 'bp_ro', false) +
            dataRow('Total RO (Mech + B&amp;P)', 'total_ro', false, 'font-weight:700;border-top:2px solid #4f6bdc;') +
            secHdr('&#128176; REVENUE', n) +
            dataRow('Labour Revenue - Service',   'service_labour',  true) +
            dataRow('Labour Revenue - Bodyshop',  'bodyshop_labour', true) +
            dataRow('Parts Retail - Mechanical',  'mech_parts',      true) +
            dataRow('Parts Retail - Body Shop',   'bp_parts',        true) +
            secHdr('&#128202; PRODUCTIVITY (Rs / RO)', n) +
            dataRow('Mechanical Parts / RO',   'mech_parts_per_ro',  true) +
            dataRow('Mechanical Labour / RO',  'mech_labour_per_ro', true) +
            dataRow('Body Repair Parts / RO',  'bp_parts_per_ro',    true) +
            dataRow('Body Repair Labour / RO', 'bp_labour_per_ro',   true);

        wrap.innerHTML =
            '<div class="op-wrap">' +
            '<table class="op-table">' +
            '<thead><tr><th class="desc-hdr">Description</th>' + colHdrs + '</tr></thead>' +
            '<tbody>' + tbody + '</tbody>' +
            '</table></div>';

    } catch(e) {
        wrap.innerHTML = '<div class="comp-loading">Error loading data - check console.</div>';
        console.error("OnePager load failed", e);
    }
}

/* ====================================================
   PAGE 5 — CURRENT MONTH (Live Google Sheet)
   ==================================================== */
function buildCmParams() {
    var p = new URLSearchParams();
    ["cm-division","cm-service","cm-sa","cm-model"].forEach(function(key) {
        (selections[key] || new Set()).forEach(function(v) { p.append(key.replace("cm-",""), v); });
    });
    return p.toString();
}

function applyCurrentMonth() { loadCurrentMonth(); }

async function loadCurrentMonth() {
    if (!window._cmJsLoaded) {
        await new Promise(function(resolve, reject) {
            var s = document.createElement("script");
            s.src = "/cm.js?v=1";
            s.onload = resolve;
            s.onerror = reject;
            document.head.appendChild(s);
        });
        window._cmJsLoaded = true;
    }
    await _loadCurrentMonth();
}

/* ====================================================
   RESET FUNCTIONS
   ==================================================== */
function resetKeys(keys) {
    keys.forEach(function(key) {
        if (selections[key]) {
            selections[key].clear();
            var listEl = document.getElementById("list-" + key);
            if (listEl) {
                listEl.querySelectorAll(".cs-item").forEach(function(item) {
                    item.classList.remove("checked");
                    item.querySelector("input").checked = false;
                    item.style.display = "";
                });
            }
            var srch = document.getElementById("search-" + key);
            if (srch) srch.value = "";
            updateFace(key);
        }
    });
}

function resetPage1() {
    document.getElementById("cy").value = "";
    resetKeys(["division","service","sa","invoice","model","month"]);
    loadCards();
    loadTable();
}

function resetPage2() {
    resetKeys(["c-division","c-service","c-sa","c-invoice","c-model"]);
    loadComparison();
}

function resetPage3() {
    resetKeys(["d-division","d-service","d-sa","d-invoice","d-model"]);
    loadDivMonth();
}

function resetPage5() {
    resetKeys(["cm-division","cm-service","cm-sa","cm-model"]);
    loadCurrentMonth();
}

function resetPage4() {
    document.getElementById("op-cy").value = "";
    resetKeys(["op-division","op-service","op-sa","op-invoice","op-model","op-month"]);
    loadOnePager();
}

/* ====================================================
   PAGE 6 — DIVISION-MONTH CY25 vs CY26
   ==================================================== */
function buildDm26Params() {
    var p = new URLSearchParams();
    ["d6-division","d6-service","d6-sa","d6-invoice","d6-model"].forEach(function(key) {
        (selections[key] || new Set()).forEach(function(v) { p.append(key.replace("d6-",""), v); });
    });
    return p.toString();
}

function applyDivMonth26() { dm26Loaded = true; loadDivMonth26(); }

function resetPage6() {
    resetKeys(["d6-division","d6-service","d6-sa","d6-invoice","d6-model"]);
    loadDivMonth26();
}

async function loadDivMonth26() {
    var wrap = document.getElementById("divmonth26-content");
    wrap.innerHTML = "<div class=comp-loading>Loading CY25 vs CY26 pivot...</div>";
    try {
        var resp = await fetch("/division-month-cy26?" + buildDm26Params());
        var data = await resp.json();
        var divisions = data.divisions;
        var months    = data.months;
        var rows      = data.rows;

        if (!divisions.length || !months.length) {
            wrap.innerHTML = "<div class=comp-loading>No CY26 data available.</div>";
            return;
        }

        var byDiv = {};
        rows.forEach(function(r) { byDiv[r.division] = r; });

        function gPct(v25, v26) {
            if (!v25) return "<span class=pct-flat>N/A</span>";
            var p = ((v26 - v25) / v25 * 100);
            var cls  = p > 0 ? "pct-up" : p < 0 ? "pct-down" : "pct-flat";
            var sign = p > 0 ? "\u25b2 +" : p < 0 ? "\u25bc " : "";
            return "<span class=" + cls + ">" + sign + p.toFixed(1) + "%</span>";
        }

        function buildPivot26(title, prefix, isCurr) {
            var f = isCurr ? function(v){ return fmt(v||0); } : function(v){ return Number(v||0).toLocaleString("en-IN"); };

            var hdr1 = "<th rowspan=2 style=min-width:80px>Division</th>";
            months.forEach(function(m) { hdr1 += "<th colspan=3>" + m + "</th>"; });
            hdr1 += "<th colspan=3>Total</th>";

            var hdr2 = "";
            months.concat(["Total"]).forEach(function() {
                hdr2 += "<th style=background:#5a7ae8;color:#fff>CY25</th>";
                hdr2 += "<th style=background:#17a34a;color:#fff>CY26</th>";
                hdr2 += "<th style=background:#2e4baa;color:#fff>Growth%</th>";
            });

            var gTot25 = new Array(months.length).fill(0);
            var gTot26 = new Array(months.length).fill(0);
            var body = "";

            divisions.forEach(function(div) {
                var r = byDiv[div];
                var cells = "";
                var rSum25 = 0, rSum26 = 0;
                months.forEach(function(m, mi) {
                    var v25 = r ? (r[prefix + "25_" + m] || 0) : 0;
                    var v26 = r ? (r[prefix + "26_" + m] || 0) : 0;
                    rSum25 += v25; rSum26 += v26;
                    gTot25[mi] += v25; gTot26[mi] += v26;
                    cells += "<td>" + f(v25) + "</td><td style=color:#15803d;font-weight:600>" + f(v26) + "</td><td>" + gPct(v25,v26) + "</td>";
                });
                body += "<tr><td>" + div + "</td>" + cells +
                    "<td style=font-weight:700>" + f(rSum25) + "</td>" +
                    "<td style=font-weight:700;color:#15803d>" + f(rSum26) + "</td>" +
                    "<td>" + gPct(rSum25, rSum26) + "</td></tr>";
            });

            var footCells = "";
            var gSum25 = 0, gSum26 = 0;
            gTot25.forEach(function(v, i) {
                gSum25 += v; gSum26 += gTot26[i];
                footCells += "<td>" + f(v) + "</td><td style=color:#15803d;font-weight:700>" + f(gTot26[i]) + "</td><td>" + gPct(v, gTot26[i]) + "</td>";
            });

            return "<div class=pivot-section comp-section>" +
                "<h3>" + title + "</h3>" +
                "<div class=pivot-wrap>" +
                "<table class=pivot-table>" +
                "<thead><tr>" + hdr1 + "</tr><tr>" + hdr2 + "</tr></thead>" +
                "<tbody>" + body + "</tbody>" +
                "<tfoot><tr><td>Grand Total</td>" + footCells +
                "<td>" + f(gSum25) + "</td>" +
                "<td style=color:#15803d;font-weight:700>" + f(gSum26) + "</td>" +
                "<td>" + gPct(gSum25, gSum26) + "</td></tr></tfoot>" +
                "</table></div></div>";
        }

        wrap.innerHTML =
            buildPivot26("Inflow - Division x Month (CY25 vs CY26)",         "i", false) +
            buildPivot26("Labour - Division x Month (CY25 vs CY26)",         "l", true)  +
            buildPivot26("Spares - Division x Month (CY25 vs CY26)",         "s", true)  +
            buildPivot26("Total Revenue - Division x Month (CY25 vs CY26)",  "t", true);

    } catch(e) {
        wrap.innerHTML = "<div class=comp-loading>Error loading data: " + e.message + "</div>";
        console.error("DivMonth26 failed:", e);
    }
}

/* ====================================================
   INIT
   ==================================================== */
window.onload = function () {
    document.querySelectorAll(".custom-select").forEach(buildCustomSelect);
    loadFilters();
    loadCards();
    loadTable();
};
/* ====================================================
   FILTER SYSTEM — COMPLETE IMPLEMENTATION
   ==================================================== */

// Division key -> [cy-selector-id, dependent-keys[]]
var filterGroups = {
    "division":    ["cy",    ["service","sa","invoice","model","month"]],
    "c-division":  [null,    ["c-service","c-sa","c-invoice","c-model"]],
    "d-division":  [null,    ["d-service","d-sa","d-invoice","d-model"]],
    "d6-division": [null,    ["d6-service","d6-sa","d6-invoice","d6-model"]],
    "op-division": ["op-cy", ["op-service","op-sa","op-invoice","op-model","op-month"]],
    "cm-division": [null,    ["cm-service","cm-sa","cm-model"]],
};

var DIV_KEYS = ["division","c-division","d-division","d6-division","op-division","cm-division"];

function isDivisionKey(key) {
    return DIV_KEYS.indexOf(key) !== -1;
}

// ── Fill dropdown WITH event listeners (initial load only) ─────────────
function fillCustomSelect(key, list) {
    var listEl = document.getElementById("list-" + key);
    if (!listEl) return;
    if (!selections[key]) selections[key] = new Set();
    listEl.innerHTML = "";
    list.forEach(function(v) {
        var item = document.createElement("div");
        item.className = "cs-item";
        item.dataset.value = v;
        var chk = document.createElement("input");
        chk.type = "checkbox"; chk.value = v;
        var lbl = document.createElement("span");
        lbl.textContent = v;
        item.appendChild(chk);
        item.appendChild(document.createTextNode(" "));
        item.appendChild(lbl);
        chk.addEventListener("change", function() {
            if (chk.checked) selections[key].add(v);
            else             selections[key].delete(v);
            item.classList.toggle("checked", chk.checked);
            updateFace(key);
            if (isDivisionKey(key)) refreshDependentFilters(key);
        });
        listEl.appendChild(item);
    });
    updateFace(key);
}

// ── Refill dropdown (dependent refresh) — keeps valid prior selections ──
function fillDropdownOnly(key, list) {
    var listEl = document.getElementById("list-" + key);
    if (!listEl) return;
    if (!selections[key]) selections[key] = new Set();

    // Keep only selections that still exist in new list
    var validVals = {};
    list.forEach(function(v) { validVals[v] = true; });
    var newSel = new Set();
    selections[key].forEach(function(v) { if (validVals[v]) newSel.add(v); });
    selections[key] = newSel;

    listEl.innerHTML = "";
    list.forEach(function(v) {
        var item = document.createElement("div");
        item.className = "cs-item";
        item.dataset.value = v;
        if (selections[key].has(v)) item.classList.add("checked");
        var chk = document.createElement("input");
        chk.type = "checkbox"; chk.value = v;
        chk.checked = selections[key].has(v);
        var lbl = document.createElement("span");
        lbl.textContent = v;
        item.appendChild(chk);
        item.appendChild(document.createTextNode(" "));
        item.appendChild(lbl);
        // Attach change listener (no cascade for dependent keys)
        chk.addEventListener("change", function() {
            if (chk.checked) selections[key].add(v);
            else             selections[key].delete(v);
            item.classList.toggle("checked", chk.checked);
            updateFace(key);
        });
        listEl.appendChild(item);
    });
    updateFace(key);
}

// ── Select All visible items ────────────────────────────────────────────
function selectAllVisible(key) {
    if (!selections[key]) selections[key] = new Set();
    var items = document.querySelectorAll("#list-" + key + " .cs-item");
    items.forEach(function(item) {
        if (item.style.display === "none") return;
        var v = item.dataset.value;
        selections[key].add(v);
        item.classList.add("checked");
        item.querySelector("input").checked = true;
    });
    updateFace(key);
    if (isDivisionKey(key)) refreshDependentFilters(key);
}

// ── Deselect All ────────────────────────────────────────────────────────
function clearSelect(key) {
    if (!selections[key]) selections[key] = new Set();
    selections[key].clear();
    var listEl = document.getElementById("list-" + key);
    if (listEl) {
        listEl.querySelectorAll(".cs-item").forEach(function(item) {
            item.classList.remove("checked");
            item.querySelector("input").checked = false;
            item.style.display = "";
        });
    }
    var srch = document.getElementById("search-" + key);
    if (srch) srch.value = "";
    updateFace(key);
    // When division cleared -> restore ALL options for dependent dropdowns
    if (isDivisionKey(key)) refreshDependentFilters(key);
}

// ── Refresh dependent dropdowns when division selection changes ─────────
async function refreshDependentFilters(divKey) {
    var group = filterGroups[divKey];
    if (!group) return;

    var cyId    = group[0];
    var depKeys = group[1];
    var divVals = [];
    if (selections[divKey]) {
        selections[divKey].forEach(function(v) { divVals.push(v); });
    }

    var cy = "";
    if (cyId) {
        var cyEl = document.getElementById(cyId);
        if (cyEl) cy = cyEl.value || "";
    }

    // If NO division selected and NO cy -> restore full options
    if (divVals.length === 0 && !cy) {
        try {
            var resp = await fetch("/filters");
            var data = await resp.json();
            var keyMap = {
                "service":"service","c-service":"service","d-service":"service",
                "d6-service":"service","op-service":"service",
                "sa":"sa","c-sa":"sa","d-sa":"sa","d6-sa":"sa","op-sa":"sa",
                "invoice":"invoice","c-invoice":"invoice","d-invoice":"invoice",
                "d6-invoice":"invoice","op-invoice":"invoice",
                "model":"model","c-model":"model","d-model":"model",
                "d6-model":"model","op-model":"model",
                "month":"month","op-month":"month",
            };
            depKeys.forEach(function(dk) {
                var dataKey = keyMap[dk];
                if (data[dataKey]) fillDropdownOnly(dk, data[dataKey]);
            });
        } catch(e) { console.warn("refreshDependentFilters (full) failed:", e); }
        return;
    }

    // Build params for filtered fetch
    var p = new URLSearchParams();
    if (cy) p.append("cy", cy);
    divVals.forEach(function(v) { p.append("division", v); });

    try {
        var resp = await fetch("/filters-dep?" + p.toString());
        var data = await resp.json();
        var keyMap = {
            "service":"service","c-service":"service","d-service":"service",
            "d6-service":"service","op-service":"service",
            "sa":"sa","c-sa":"sa","d-sa":"sa","d6-sa":"sa","op-sa":"sa",
            "invoice":"invoice","c-invoice":"invoice","d-invoice":"invoice",
            "d6-invoice":"invoice","op-invoice":"invoice",
            "model":"model","c-model":"model","d-model":"model",
            "d6-model":"model","op-model":"model",
            "month":"month","op-month":"month",
        };
        depKeys.forEach(function(dk) {
            var dataKey = keyMap[dk];
            if (data[dataKey]) fillDropdownOnly(dk, data[dataKey]);
        });
    } catch(e) {
        console.warn("refreshDependentFilters failed:", e.message);
    }
}


</script>
</body>
</html>""")
# ---------------- AUTO OPEN ----------------

def open_browser():
    webbrowser.open("http://127.0.0.1:8000")


if __name__ == "__main__":
    threading.Timer(1, open_browser).start()
    uvicorn.run(app, host="127.0.0.1", port=8000)
